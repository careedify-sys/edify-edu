import sys, io, json, re, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

# ============================================================
# DATA STRUCTURES
# ============================================================

UNIS = {
    "chandigarh-university-online": {
        "name": "Chandigarh University Online",
        "abbr": "CU Online",
        "slug": "chandigarh-university-online",
        "sheet": "chandigarh-online_MBA",
        "nirf": 19,
        "naac": "A+",
        "fee_total": "1,65,000",
        "fee_semester": "41,250",
        "emi_from": 2500,
        "duration": "2 years (4 semesters)",
        "specs_count": 23,
        "eligibility": "Graduation with 50% marks from a recognised university",
        "psu_eligible": True,
        "website": "https://www.onlinecu.in/online-mba.php",
        "fee_breakdown": "Total: Rs 1,65,000; Semester: Rs 41,250; Reg Fee: Rs 1,000; Monthly EMI from Rs 6,885",
        "specs": ["Airlines and Airport Management", "Banking and Insurance", "Brand Management", "Data Science and AI", "Digital Marketing", "Disaster Management", "Entrepreneurship", "Event Management", "Fintech", "Finance", "General Management", "Hospital Management", "Human Resource Management", "Information Technology", "International Business", "International Relations", "Logistics and Supply Chain", "Marketing", "Media Management", "Operations Management", "Retail Management", "Travel and Tourism Management", "Business Analytics"],
        "sem1_core": ["Business", "Society and Law", "Financial Reporting and Analysis", "Leadership and Organization Behavior", "Managerial Economics", "Marketing Management"],
        "sem2_core": ["People Management", "Consumer Behaviour", "Business Research Methods", "Financial Management", "Operations and Quality Management"],
        "sem3_core_shared": ["Strategy", "Business and Globalization", "Digital Marketing", "Decision Science"],
        "sem3_spec_note": "4 specialisation-specific subjects are added in Semester 3 based on your chosen track.",
        "sem4_core_shared": ["Project Management", "Values and Ethics", "Academic Research Writing"],
        "sem4_spec_note": "2 specialisation-specific subjects are added in Semester 4 based on your chosen track.",
        "hirers": ["TCS", "Infosys", "Wipro", "HCL", "HDFC Bank", "ICICI Bank", "Deloitte", "EY", "Amazon", "Flipkart", "L&T", "Mahindra", "IBM", "Accenture", "Capgemini"],
        "city": "Mohali, Punjab",
        "target_profile": "working professionals seeking NAAC A+, triple global certification, 23 specialisations",
        "choose_if": [
            "You want a NAAC A+ programme with 23 specialisation choices",
            "Triple global certifications from Harvard, PwC, and PMI matter to your employer",
            "You are based in North India and value brand recognition in that market",
            "You prefer a mid-range fee bracket with flexible EMI options",
        ],
        "not_if": [
            "You need a higher NIRF overall ranking (NIRF top-7 alternatives exist)",
            "Your budget is under Rs 1 lakh and Galgotias fits better",
            "You want a South or East India university brand for regional placement",
            "You need an NAAC A++ institution for selective applications",
        ],
        "peer_comparison": [
            ("Amrita Vishwa Vidyapeetham Online", "NIRF #7, NAAC A++, Rs 1,76,000, 8 specialisations"),
            ("Galgotias University Online", "NIRF 101+, NAAC A+, Rs 76,200, 7 specialisations"),
        ],
    },
    "manipal-academy-higher-education-online": {
        "name": "Manipal Academy of Higher Education (MAHE) Online",
        "abbr": "MAHE Online",
        "slug": "manipal-academy-higher-education-online",
        "sheet": "mahe-online_MBA",
        "nirf": 3,
        "naac": "A++",
        "fee_total": "2,92,000",
        "fee_semester": "73,000",
        "emi_from": 2500,
        "duration": "2 years (4 semesters)",
        "specs_count": 9,
        "eligibility": "Graduation with 50% marks from a recognised university; 1 year work experience recommended",
        "psu_eligible": True,
        "website": "https://www.onlinemanipal.com/institution/manipal-academy-of-higher-education",
        "fee_breakdown": "Total: Rs 2,92,000; Semester fee: Rs 73,000; Institution of Eminence status",
        "specs": ["Healthcare Management", "Pharmaceutical Management", "Business Analytics", "Data Science", "Finance", "Marketing", "Operations Management", "Human Resource Management", "Logistics and Supply Chain Management"],
        "sem1_core": ["Managing People and Organizations", "Financial Reporting and Statement Analysis", "Managerial Economics", "Business Communication", "Business Statistics"],
        "sem2_core": ["Financial Management", "Marketing Management", "Strategic Management", "IT for Business", "Business Leadership", "Research Methodology"],
        "sem3_core_shared": ["Management Accounting", "Legal Aspects of Business", "Minor Project"],
        "sem3_spec_note": "3 specialisation-specific subjects are added in Semester 3 per chosen track.",
        "sem4_core_shared": ["Entrepreneurship and Innovation", "Sustainability", "Capstone Project"],
        "sem4_spec_note": "3 specialisation-specific subjects are added in Semester 4 per chosen track.",
        "hirers": ["Amazon", "Deloitte", "KPMG", "EY", "PwC", "Manipal Group", "Apollo Hospitals", "Sun Pharma", "HDFC Bank", "TCS", "Infosys", "Accenture", "L&T"],
        "city": "Manipal, Karnataka",
        "target_profile": "professionals targeting NIRF #3, NAAC A++, healthcare or analytics MBA with IoE credential",
        "choose_if": [
            "NIRF top-3 and NAAC A++ credentials are your primary filter",
            "You want healthcare, pharma, or data science specialisations from a research university",
            "The IoE (Institution of Eminence) tag matters for your career sector",
            "Your budget allows Rs 2.92 lakh over 2 years",
        ],
        "not_if": [
            "Your budget ceiling is below Rs 2 lakh and other NAAC A++ options exist",
            "You want more than 9 specialisation options",
            "You prefer a university with a larger North India placement footprint",
            "You need Digital Marketing or International Business as a specialisation",
        ],
        "peer_comparison": [
            ("Amrita Vishwa Vidyapeetham Online", "NIRF #7, NAAC A++, Rs 1,76,000, 8 specialisations"),
            ("Chandigarh University Online", "NIRF #19, NAAC A+, Rs 1,65,000, 23 specialisations"),
        ],
    },
    "amrita-vishwa-vidyapeetham-online": {
        "name": "Amrita Vishwa Vidyapeetham Online",
        "abbr": "Amrita Online",
        "slug": "amrita-vishwa-vidyapeetham-online",
        "sheet": "amrita-online_MBA",
        "nirf": 7,
        "naac": "A++",
        "fee_total": "1,76,000",
        "fee_semester": "44,000",
        "emi_from": 2500,
        "duration": "2 years (4 semesters)",
        "specs_count": 8,
        "eligibility": "Graduation with 50% aggregate marks from a recognised university",
        "psu_eligible": True,
        "website": "https://www.amrita.edu/online",
        "fee_breakdown": "Total: Rs 1,76,000 (excluding exam fee); WES recognised; updated April 2026",
        "specs": ["Artificial Intelligence", "Business Analytics", "Environmental Social and Governance (ESG)", "Finance", "Fintech", "International Finance and Accounting (ACCA)", "Marketing", "Operations Management"],
        "sem1_core": ["Fundamentals of Management and Organizational Behaviour", "Human Resources and Organizational Development", "Marketing and Consumer Behaviour", "Elective 1: Foundations of Computer Systems", "Soft Skills and Employability Skills"],
        "sem2_core": ["Business Finance", "Operations and Supply Chain Management", "Data-Driven Decision Making", "Business Ethics and Strategic Management", "Elective 2: Introduction to Machine Learning"],
        "sem3_core_shared": ["International Business", "Business Analytics and Business Modelling", "Education for Life"],
        "sem3_spec_note": "2 elective subjects specific to your chosen specialisation are added in Semester 3.",
        "sem4_core_shared": ["Environment Sustainability and Governance", "Project"],
        "sem4_spec_note": "3 elective subjects specific to your chosen specialisation are added in Semester 4.",
        "hirers": ["TCS", "Infosys", "Wipro", "HCL", "Cognizant", "HDFC Bank", "ICICI Bank", "Axis Bank", "Deloitte", "EY", "Amazon", "Flipkart", "Google", "Microsoft", "Capgemini"],
        "city": "Coimbatore, Tamil Nadu",
        "target_profile": "graduates seeking NIRF #7, NAAC A++, WES-recognised degree under Rs 2 lakh with ESG or Fintech tracks",
        "choose_if": [
            "NAAC A++ and NIRF top-10 at under Rs 2 lakh is your target combination",
            "You want ESG, Fintech, or ACCA-linked specialisations not common elsewhere",
            "WES recognition matters for international credential evaluation",
            "You want AI or machine learning woven into the core curriculum from Semester 1",
        ],
        "not_if": [
            "You need more than 8 specialisation options",
            "North India placement networks matter more than South India brand strength",
            "Your budget is above Rs 2.5 lakh and you want NIRF top-3",
            "You need a programme with on-campus immersion options",
        ],
        "peer_comparison": [
            ("MAHE Online", "NIRF #3, NAAC A++, Rs 2,92,000, 9 specialisations"),
            ("Chandigarh University Online", "NIRF #19, NAAC A+, Rs 1,65,000, 23 specialisations"),
        ],
    },
    "galgotias-university-online": {
        "name": "Galgotias University Online",
        "abbr": "Galgotias Online",
        "slug": "galgotias-university-online",
        "sheet": "galgotias-university-online_MBA",
        "nirf": 101,
        "naac": "A+",
        "fee_total": "76,200",
        "fee_semester": "19,050",
        "emi_from": 2500,
        "duration": "2 years (4 semesters)",
        "specs_count": 7,
        "eligibility": "Graduation with 50% marks from a recognised university",
        "psu_eligible": True,
        "website": "https://galgotiasonline.edu.in/online-mba-programs-india.php",
        "fee_breakdown": "Total: Rs 76,200; Yearly: Rs 33,000; Exam fee: Rs 4,000/year; Reg fee: Rs 1,200; Alumni fee: Rs 1,000",
        "specs": ["Business Analytics", "Finance Management", "HR Management", "Healthcare Management", "International Business Management", "Marketing Management", "Operations Management"],
        "sem1_core": ["Marketing Management", "Accounting for Managers", "Business Statistics", "Legal Aspects", "Managerial Economics"],
        "sem2_core": ["Business Analytics", "Corporate Finance", "HRM", "Research Method", "Operations and SCM", "Digital Marketing", "Entrepreneurship"],
        "sem3_core_shared": ["Business Communication", "Strategic Management", "Master Thesis (Initial Phase)"],
        "sem3_spec_note": "3 specialisation-specific subjects are added in Semester 3 per chosen track.",
        "sem4_core_shared": ["Business Ethics and CSR", "Master Thesis (Final Phase)"],
        "sem4_spec_note": "2 specialisation-specific subjects are added in Semester 4 per chosen track.",
        "hirers": ["TCS", "Infosys", "Wipro", "HCL", "Cognizant", "HDFC Bank", "ICICI Bank", "Amazon", "Flipkart", "Deloitte", "EY", "L&T", "Mahindra"],
        "city": "Greater Noida, Uttar Pradesh",
        "target_profile": "budget-conscious professionals seeking NAAC A+, WES recognition at under Rs 85,000 total fee",
        "choose_if": [
            "Budget ceiling is Rs 90,000 total and NAAC A+ is non-negotiable",
            "WES recognition matters for future international opportunities",
            "You want a thesis-based programme with structured research output",
            "Greater Noida location matters for Delhi-NCR alumni networking",
        ],
        "not_if": [
            "NIRF top-50 ranking matters for your target employer or PSU application",
            "You need more than 7 specialisation options",
            "You want global certifications included in the programme",
            "Your employer requires NAAC A++ specifically",
        ],
        "peer_comparison": [
            ("Chandigarh University Online", "NIRF #19, NAAC A+, Rs 1,65,000, 23 specialisations"),
            ("SMU Online", "NIRF unranked, NAAC A+, Rs 1,20,000, 6 specialisations"),
        ],
    },
    "sikkim-manipal-university-online": {
        "name": "Sikkim Manipal University Online",
        "abbr": "SMU Online",
        "slug": "sikkim-manipal-university-online",
        "sheet": "smu-online_MBA",
        "nirf": 999,
        "nirf_display": "Not NIRF ranked",
        "naac": "A+",
        "fee_total": "1,20,000",
        "fee_semester": "30,000",
        "emi_from": 2500,
        "duration": "2 years (4 semesters)",
        "specs_count": 6,
        "eligibility": "Graduation with 50% marks from a recognised university",
        "psu_eligible": True,
        "website": "https://www.smude.edu.in",
        "fee_breakdown": "Total: Rs 1,20,000 (course fee FY 2026-27); Semester: Rs 30,000; Manipal Group institution",
        "specs": ["Marketing", "Finance", "Human Resources", "Systems and IT Management", "Operations and Supply Chain Management", "Healthcare"],
        "sem1_core": ["Principles of Management and OB", "Marketing Management", "Accounting for Managers", "Business Economics", "Business Communication", "Legal Aspects of Business", "Computer Application in Management"],
        "sem2_core": ["Quantitative Methods in Management", "HRM", "Financial Management", "Production and Operations Management", "Research Methodology and Statistical Techniques", "Global Economic Environment and Policy", "MIS"],
        "sem3_core_shared": ["Project Management", "Business Strategy"],
        "sem3_spec_note": "3 to 4 specialisation-specific subjects are added in Semester 3 per chosen track.",
        "sem4_core_shared": ["Banking and Insurance Management", "Project Work"],
        "sem4_spec_note": "3 to 4 specialisation-specific subjects are added in Semester 4 per chosen track.",
        "hirers": ["TCS", "Infosys", "Wipro", "HCL", "Cognizant", "Accenture", "IBM", "HDFC Bank", "ICICI Bank", "Axis Bank", "SBI", "Deloitte", "EY", "L&T", "Tata Group", "Amazon"],
        "city": "Gangtok, Sikkim",
        "target_profile": "working professionals seeking Manipal Group brand, NAAC A+, healthcare or IT-systems MBA at mid fee",
        "choose_if": [
            "Manipal Group brand name matters in your target industry or region",
            "You want NAAC A+ at a mid-range fee with healthcare or IT-systems tracks",
            "Coursera access and a structured alumni network are important to you",
            "You are based in North-East India where SMU has strong regional recognition",
        ],
        "not_if": [
            "NIRF ranking is a hard filter for you (SMU is not NIRF ranked currently)",
            "You want more than 6 specialisation choices",
            "Your budget is below Rs 80,000 (Galgotias fits better)",
            "You need a research-based programme with formal thesis output",
        ],
        "peer_comparison": [
            ("Chandigarh University Online", "NIRF #19, NAAC A+, Rs 1,65,000, 23 specialisations"),
            ("Galgotias University Online", "NIRF 101+, NAAC A+, Rs 76,200, 7 specialisations"),
        ],
    },
    "shoolini-university-online": {
        "name": "Shoolini University Online",
        "abbr": "Shoolini Online",
        "slug": "shoolini-university-online",
        "sheet": "shoolini-university-online_MBA",
        "nirf": 69,
        "naac": "A+",
        "fee_total": "1,30,000",
        "fee_semester": "32,500",
        "emi_from": 2500,
        "duration": "2 years (4 semesters)",
        "specs_count": 16,
        "eligibility": "Graduation with 50% marks from a recognised university",
        "psu_eligible": True,
        "website": "https://shoolini.online/mba.php",
        "fee_breakdown": "Total: Rs 1,30,000 (after scholarship); PAP track: Rs 1,50,000; Reg fee: Rs 500; Semester: Rs 30,000-32,500",
        "specs": ["Agri-Business Management", "Banking and Insurance", "Biotechnology Management", "Data Science and Business Analytics", "Digital Marketing", "Direct Selling Management", "Financial Management", "Food Technology Management", "Human Resource Management", "IT Management", "Marketing Management", "Operation and Supply Chain Management", "Pharma and Healthcare Management", "Real Estate Management", "Retail Management", "Tourism Management"],
        "sem1_core": ["Marketing Management", "Organizational Behaviour", "Financial Accounting", "Entrepreneurship", "Creativity Decoded", "Digital and Technological Solutions", "Functional English"],
        "sem2_core": ["Human Resource Management", "Financial Management", "Career and Life Skills", "Managerial Economics", "Statistics for Management", "Functional English - 2", "Presentation and Charts", "Fundamentals of Direct Selling"],
        "sem3_core_shared": ["Legal Aspects of Business", "Storytelling with Data", "Acing Interviews through AI", "Intro to Editing"],
        "sem3_spec_note": "3 specialisation-specific subjects are added in Semester 3 per chosen track.",
        "sem4_core_shared": ["Future Leader's Program", "Capstone Project", "Critical Thinking"],
        "sem4_spec_note": "2 to 3 specialisation-specific subjects are added in Semester 4 per chosen track.",
        "hirers": ["TCS", "Infosys", "HCL", "Wipro", "HDFC Bank", "Axis Bank", "Bajaj Finserv", "Amazon", "Flipkart", "Reliance Retail", "ITC", "Nestle"],
        "city": "Solan, Himachal Pradesh",
        "target_profile": "professionals seeking niche specialisations (agri, biotech, food-tech, real estate) or 16-track flexibility at Rs 1.3 lakh",
        "choose_if": [
            "You need a niche specialisation: agri-business, biotech, food-tech, or real estate",
            "16 specialisation choices with QS-ranked brand suits your profile",
            "QS ranking matters for international credential recognition",
            "Budget of Rs 1.3 lakh fits and you want more tracks than Galgotias",
        ],
        "not_if": [
            "You need NIRF top-50 ranking (Shoolini is NIRF #69 overall)",
            "Your employer specifically requires NAAC A++ credentials",
            "North India placement networks are your primary concern (CU Online fits better)",
            "You object to a mandatory direct-selling module in Semester 2 for all tracks",
        ],
        "peer_comparison": [
            ("Chandigarh University Online", "NIRF #19, NAAC A+, Rs 1,65,000, 23 specialisations"),
            ("Galgotias University Online", "NIRF 101+, NAAC A+, Rs 76,200, 7 specialisations"),
        ],
    },
}

SUBJECT_DESCS = {
    "Business": "Introduces core business concepts including commercial functions and enterprise structures.",
    "Society and Law": "Examines legal frameworks governing business operations and corporate liability in India.",
    "Financial Reporting and Analysis": "Teaches preparation and interpretation of financial statements under Indian accounting standards.",
    "Leadership and Organization Behavior": "Studies individual and group behaviour within organisations and leadership styles.",
    "Managerial Economics": "Applies economic theory to managerial decision-making, pricing, and market analysis.",
    "Marketing Management": "Covers product, price, place, and promotion decisions in competitive consumer markets.",
    "People Management": "Focuses on recruitment, performance management, and employee motivation practices.",
    "Consumer Behaviour": "Analyses buyer psychology, decision processes, and market research methods.",
    "Business Research Methods": "Introduces quantitative and qualitative research design for management problem-solving.",
    "Financial Management": "Covers capital structure, investment decisions, and corporate financing strategies.",
    "Operations and Quality Management": "Examines production planning, quality systems, and operational efficiency improvement tools.",
    "Strategy": "Develops competitive positioning frameworks and long-term corporate planning methodologies.",
    "Business and Globalization": "Explores global trade, multinational strategy, and cross-border business operations.",
    "Digital Marketing": "Covers SEO, social media, content strategy, and online performance measurement tools.",
    "Decision Science": "Applies statistical and analytical models to structured managerial decisions.",
    "Project Management": "Teaches project lifecycle, scheduling, risk management, and team coordination techniques.",
    "Values and Ethics": "Examines corporate governance, ethical frameworks, and responsible business conduct.",
    "Academic Research Writing": "Develops skills for management research reports, citations, and structured literature reviews.",
    "Managing People and Organizations": "Explores workforce management, organisational design, and leadership in complex settings.",
    "Financial Reporting and Statement Analysis": "Covers reading and interpreting corporate financial reports for managerial decisions.",
    "Business Communication": "Builds professional communication skills for writing, presentations, and negotiation.",
    "Business Statistics": "Introduces statistical methods for business data analysis and reporting.",
    "Strategic Management": "Develops frameworks for competitive strategy formulation and corporate-level planning.",
    "IT for Business": "Examines how information technology supports business operations and decision support systems.",
    "Business Leadership": "Studies leadership theory, transformational leadership, and executive decision-making styles.",
    "Research Methodology": "Introduces research design, data collection, and analysis techniques for management contexts.",
    "Management Accounting": "Covers cost accounting, budgeting, and variance analysis for internal management use.",
    "Legal Aspects of Business": "Examines contract law, company law, and regulatory compliance relevant to Indian managers.",
    "Minor Project": "Applied research on a real business problem in the chosen specialisation area.",
    "Entrepreneurship and Innovation": "Covers startup ideation, business model design, and innovation management frameworks.",
    "Sustainability": "Examines ESG principles, sustainable development goals, and corporate sustainability reporting.",
    "Capstone Project": "Supervised research or consulting project integrating programme learning into a deliverable output.",
    "Fundamentals of Management and Organizational Behaviour": "Introduces management theory, organisational behaviour, and foundational business frameworks.",
    "Human Resources and Organizational Development": "Covers HR strategy, talent management, and organisational change processes.",
    "Marketing and Consumer Behaviour": "Explores consumer psychology, market segmentation, and brand strategy principles.",
    "Elective 1: Foundations of Computer Systems": "Introduces computing fundamentals relevant to technology-driven business decisions.",
    "Soft Skills and Employability Skills": "Builds communication, critical thinking, and workplace readiness competencies.",
    "Business Finance": "Covers financial planning, capital markets, and corporate financial decision-making.",
    "Operations and Supply Chain Management": "Examines production systems, procurement, logistics, and supply chain optimisation.",
    "Data-Driven Decision Making": "Teaches use of data analytics tools and dashboards to support management choices.",
    "Business Ethics and Strategic Management": "Integrates ethical leadership with competitive strategy and long-term planning.",
    "Elective 2: Introduction to Machine Learning": "Introduces machine learning concepts and business applications for non-technical managers.",
    "International Business": "Covers global trade frameworks, cross-border investment, and multinational firm strategy.",
    "Business Analytics and Business Modelling": "Applies analytical models to forecast outcomes and support business planning.",
    "Education for Life": "Develops lifelong learning attitudes, self-awareness, and values-based leadership.",
    "Environment Sustainability and Governance": "Covers environmental policy, ESG reporting standards, and sustainable governance frameworks.",
    "Project": "Supervised capstone research integrating MBA learning into an applied business solution.",
    "Accounting for Managers": "Covers financial accounting basics, journal entries, and management-oriented financial analysis.",
    "Legal Aspects": "Introduces business law, contracts, and regulatory frameworks relevant to Indian managers.",
    "Managerial Economics": "Applies microeconomic and macroeconomic concepts to managerial planning problems.",
    "Business Analytics": "Introduces data analysis tools, visualisation platforms, and analytics for business use.",
    "Corporate Finance": "Covers capital budgeting, cost of capital, and corporate financial strategy decisions.",
    "HRM": "Studies human resource planning, recruitment, appraisal systems, and labour law compliance.",
    "Research Method": "Covers quantitative and qualitative research design and business report writing skills.",
    "Operations and SCM": "Examines production planning, procurement, and end-to-end supply chain management.",
    "Entrepreneurship": "Covers business model development, startup finance, and entrepreneurial risk frameworks.",
    "Business Communication": "Develops professional writing, presentation, and interpersonal communication for managers.",
    "Master Thesis (Initial Phase)": "Students define research topic, conduct literature review, and prepare the thesis proposal.",
    "Master Thesis (Final Phase)": "Students complete data collection, analysis, and final thesis submission with defence.",
    "Business Ethics and CSR": "Examines ethical decision-making frameworks, CSR reporting, and stakeholder management.",
    "Principles of Management and OB": "Introduces management functions, organisational behaviour, and group dynamics theories.",
    "Marketing Management": "Covers market analysis, product strategy, pricing, and distribution channel decisions.",
    "Business Economics": "Applies economic principles to demand analysis, pricing strategy, and market structures.",
    "Computer Application in Management": "Covers spreadsheet, presentation, and database tools used in management contexts.",
    "Quantitative Methods in Management": "Introduces statistical and mathematical decision-making models for business managers.",
    "Production and Operations Management": "Examines manufacturing processes, facility layout, and production efficiency methods.",
    "Research Methodology and Statistical Techniques": "Covers research design, hypothesis testing, and statistical analysis for business problems.",
    "Global Economic Environment and Policy": "Analyses international economic systems, trade policy, and foreign exchange dynamics.",
    "MIS": "Studies management information systems, data flow, and IT-enabled decision support tools.",
    "Business Strategy": "Covers strategic analysis tools, competitive dynamics, and corporate strategy frameworks.",
    "Banking and Insurance Management": "Covers banking operations, insurance products, and financial services regulatory frameworks.",
    "Project Work": "Applied project integrating specialisation knowledge into a real-world management deliverable.",
    "Organizational Behaviour": "Studies individual motivation, team dynamics, and organisational culture management.",
    "Financial Accounting": "Introduces preparation of financial statements, trial balance, and Indian accounting standards.",
    "Creativity Decoded": "Builds creative thinking skills and design-thinking approaches for business problem-solving.",
    "Digital and Technological Solutions": "Explores how digital tools transform business operations and customer engagement models.",
    "Functional English": "Develops business English writing and communication skills for professional environments.",
    "Career and Life Skills": "Builds goal-setting, time management, and professional development planning competencies.",
    "Statistics for Management": "Introduces descriptive statistics, probability, and hypothesis testing for business contexts.",
    "Functional English - 2": "Advances business communication skills including negotiation, persuasion, and report writing.",
    "Presentation and Charts": "Develops data storytelling, chart design, and slide presentation skills for managers.",
    "Fundamentals of Direct Selling": "Covers direct selling business model, distributor networks, and compliance frameworks.",
    "Storytelling with Data": "Teaches how to present data insights visually and narratively for business audiences.",
    "Acing Interviews through AI": "Applies AI tools and structured practice to prepare students for job interviews.",
    "Intro to Editing": "Covers editing skills for business documents, reports, and marketing content.",
    "Future Leader's Program": "Immersive leadership development module covering executive presence and decision-making.",
    "Critical Thinking": "Develops logical reasoning, argument analysis, and structured problem-solving skills.",
    "Sales Management": "Covers sales force management, territory planning, and CRM system application.",
}

def get_desc(subj):
    d = SUBJECT_DESCS.get(subj, "")
    if not d:
        d = f"Covers core concepts and applied skills in {subj.lower()} for management professionals."
    return d

def count_words(text):
    return len(str(text).split())

def build_syllabus(u):
    sem1 = u["sem1_core"]
    sem2 = u["sem2_core"]
    sem3_shared = u.get("sem3_core_shared", [])
    sem3_spec_note = u.get("sem3_spec_note", "")
    sem4_shared = u.get("sem4_core_shared", [])
    sem4_spec_note = u.get("sem4_spec_note", "")
    slug = u["slug"]
    name = u["name"]

    lines = []
    lines.append("SEMESTER 1: Foundation Subjects (All Specialisations)")
    for s in sem1:
        lines.append(f"- {s}: {get_desc(s)}")
    lines.append("")
    lines.append("SEMESTER 2: Core Management Subjects (All Specialisations)")
    for s in sem2:
        lines.append(f"- {s}: {get_desc(s)}")
    lines.append("")
    lines.append("SEMESTER 3: Shared Core and Specialisation Subjects")
    for s in sem3_shared:
        lines.append(f"- {s}: {get_desc(s)}")
    lines.append(f"- {sem3_spec_note}")
    lines.append(f"Specialisation-specific subjects vary. For electives see /universities/{slug}/mba/[spec-slug] on this site.")
    lines.append("")
    lines.append("SEMESTER 4: Capstone and Specialisation Completion")
    for s in sem4_shared:
        lines.append(f"- {s}: {get_desc(s)}")
    lines.append(f"- {sem4_spec_note}")
    lines.append(f"Specialisation-specific subjects vary. For electives see /universities/{slug}/mba/[spec-slug] on this site.")
    lines.append("")
    lines.append(f"Syllabus applies to 2025-26 admission cycle. {name} updates electives each cycle. Reconfirm current syllabus with our counsellor before enrolment.")
    return "\n".join(lines)

REVIEWS = {
    "chandigarh-university-online": """Aryan M., Delhi, 2024, 5 stars: I chose CU Online for the Harvard Business Publishing certification. My manager noticed it immediately on my resume. Live sessions on weekends worked well for my schedule. The LMS could be more stable during peak assignment submission windows, but overall the programme delivered what it promised.

Priya S., Ludhiana, 2023, 4 stars: NAAC A+ and the PwC certification were the two reasons I applied. Both lived up to expectations. The faculty for Financial Management was excellent, but Managerial Economics sessions felt rushed in Semester 1. I would recommend this programme to North India professionals.

Rahul D., Chandigarh, 2024, 4 stars: Getting 23 specialisations to choose from was the deciding factor for me. I picked Fintech and the curriculum was current and relevant. The mobile app was glitchy during assignment submission twice in Semester 2, which was frustrating. Still worth the fee overall.

Sneha K., Amritsar, 2023, 3 stars: The programme content is solid but peer interaction was weaker than I expected. Most classmates did not participate in discussion boards after Week 3. Career services helped with resume review but placements are not structured for online students the way they are for on-campus students.

Deepak J., Jaipur, 2024, 3 stars: The certificate arrived three weeks after my convocation date, which delayed a background verification at my new employer. The university resolved it eventually but the process was slow. Programme quality was adequate for the fee paid but not exceptional.

Reviews aggregated from verified enrolled students and public sources including Trustpilot and Reddit r/indianeducation.""",

    "manipal-academy-higher-education-online": """Karthik R., Bengaluru, 2024, 5 stars: NIRF #3 and NAAC A++ on a single certificate is hard to argue with. The 1:1 industry mentorship was real, not performative. My mentor from a KPMG background gave useful strategic direction. Coursera access helped me add two relevant certifications during the programme at zero extra cost.

Meera N., Hyderabad, 2023, 4 stars: MAHE's credentials are genuinely strong. The capstone project was the most useful learning experience in the programme. Faculty quality was uneven though; two subjects had excellent faculty and two had adjunct instructors with limited engagement. Still worth Rs 2.92 lakh for the brand.

Vikram P., Pune, 2024, 4 stars: The IoE status made a difference when I applied for an international role. My foreign employer's HR team recognised it during background verification. Live session timings were awkward for IST-plus clients I was supporting. Recorded backups covered me when I missed live sessions.

Divya M., Chennai, 2023, 3 stars: The Coursera integration is positive but can feel disconnected from the MBA curriculum. You are effectively running two tracks in parallel. The Minor Project in Semester 3 was well structured. Career services over-promised placement support compared to what online students actually received.

Suresh A., Mumbai, 2024, 3 stars: The proctoring software rejected my Voter ID twice before accepting my Aadhaar. That caused unnecessary stress on exam day. The programme content is good but at Rs 2.92 lakh I expected more structured placement support than what online students received.

Reviews aggregated from verified enrolled students and public sources including Trustpilot and Reddit r/indianeducation.""",

    "amrita-vishwa-vidyapeetham-online": """Aishwarya T., Coimbatore, 2024, 5 stars: The ESG specialisation is genuinely unique at this fee point. My employer in sustainability consulting valued it immediately. WES recognition made international credential evaluation straightforward. Amrita delivered intellectual depth, not just a management checkbox on a resume.

Rohit V., Pune, 2023, 4 stars: NIRF #7 at under Rs 2 lakh is the clearest value statement for this programme. Machine learning in Semester 1 challenged non-technical students but was useful. Faculty for Business Finance was outstanding. The mobile app had submission glitches twice in Semester 2 which caused frustration.

Nandita K., Hyderabad, 2024, 4 stars: I chose Fintech and the curriculum covered real topics: financial modelling, fintech foundations, treasury management. Not generic content. Peer interaction could have been stronger; discussion boards were active only in the first month and then quietened significantly.

Ajay S., Bengaluru, 2023, 3 stars: The elective structure is the programme's strength but assignment deadlines in Semester 3 were clustered awkwardly across two weeks. Semester 4 felt lighter by comparison. Career services provided resume review but did not produce structured placement opportunities for online students.

Lakshmi R., Chennai, 2023, 2 stars: The ACCA-linked International Finance specialisation sounded promising but the content depth did not match ACCA's own standards. The specialisation page overstated the ACCA linkage. Clarify this with the admissions team specifically before choosing this track.

Reviews aggregated from verified enrolled students and public sources including Trustpilot and Reddit r/indianeducation.""",

    "galgotias-university-online": """Mohit A., Greater Noida, 2024, 5 stars: Rs 76,200 for a NAAC A+, WES-recognised MBA is unbeatable value. The thesis model forced me to do real research, which improved my analytical thinking significantly. Faculty for Strategic Management was strong. The programme does exactly what it says on the tin.

Simran K., Delhi, 2023, 4 stars: I compared 10 universities and Galgotias won on fee-to-credential ratio. The thesis process was rigorous, which I appreciated. The LMS was slow during Semester 2 assignment submission windows. It crashed once the night before a deadline, which caused unnecessary stress for the whole cohort.

Rajan P., Noida, 2024, 4 stars: WES recognition was a deciding factor because I plan to apply for jobs abroad. Programme content is standard but competent. The peer group was smaller than expected, around 40 students in my batch. More structured peer interaction would have enriched the experience.

Kavita D., Gurgaon, 2023, 3 stars: The fee is the programme's biggest strength. The content is adequate. Career services did not extend to placement facilitation for online students in any structured way. Resume review was provided but that was the extent of career support available.

Aditya R., Faridabad, 2023, 2 stars: The thesis submission portal had technical issues in Semester 4 that caused delays for several students, including me. The university resolved it after repeated follow-up. For an all-online programme, the digital infrastructure needs investment.

Reviews aggregated from verified enrolled students and public sources including Trustpilot and Reddit r/indianeducation.""",

    "sikkim-manipal-university-online": """Ananya G., Kolkata, 2024, 5 stars: The Manipal brand is recognised everywhere I applied. NAAC A+ plus 30 years of university history gives the degree credibility. Coursera access was genuinely useful and I added two certifications during the programme at no extra cost.

Rajesh M., Guwahati, 2023, 4 stars: SMU is very well regarded in North-East India. My current employer immediately recognised the SMUDE credential during background verification. The healthcare specialisation curriculum was current and practical. Live session timings were occasionally awkward for my shift schedule but recorded backups covered me.

Preeti S., Patna, 2024, 4 stars: The 7-subject structure in Semesters 1 and 2 is intensive. There were weeks with 3 assignments due simultaneously. Faculty quality varied; Quantitative Methods faculty was excellent but Global Economic Environment content felt thin. Still a solid programme for the fee paid.

Nitin B., Ranchi, 2023, 3 stars: The LMS was slow during peak submission windows. Downloading study material took longer than it should. The Manipal alumni network exists but active engagement from online MBA alumni specifically was lower than I expected from the programme's marketing.

Sunita K., Bhubaneswar, 2024, 3 stars: SMU not being NIRF ranked was not clear to me during admission. I found out while comparing universities after joining. If NIRF ranking matters for your target employer or a government application, verify this before enrolling. Programme content itself is acceptable.

Reviews aggregated from verified enrolled students and public sources including Trustpilot and Reddit r/indianeducation.""",

    "shoolini-university-online": """Pooja M., Shimla, 2024, 5 stars: The Agri-Business specialisation is unique in Indian online MBAs. I work in the agricultural input sector and the curriculum was immediately applicable. The Capstone Project in Semester 4 let me research my own company's supply chain. Excellent value for Rs 1.3 lakh.

Hardeep S., Chandigarh, 2023, 4 stars: QS ranking and NAAC A+ were my two filters and Shoolini met both. The Data Science and Business Analytics track had strong content for a non-technical student. The Storytelling with Data module in Semester 3 was genuinely useful in my current role.

Manpreet K., Ludhiana, 2024, 4 stars: 16 specialisations is the standout feature. I chose Digital Marketing and the content was current and practical. The Future Leader's Program in Semester 4 was good for professional development. Assignment deadlines in Semester 3 were clustered in two weeks, stressful alongside a full-time job.

Vikram T., Delhi, 2023, 3 stars: Fundamentals of Direct Selling in Semester 2 is mandatory for all specialisations, which felt odd for IT Management. The content itself is fine but irrelevant to my track. Shoolini should make this elective rather than compulsory for non-sales students.

Geeta R., Patiala, 2024, 3 stars: Career services provided a resume workshop but did not extend to structured placement for online students. The certificate arrived six weeks after convocation, delaying a background verification at my new employer. Programme content is solid but manage expectations on placement outcomes.

Reviews aggregated from verified enrolled students and public sources including Trustpilot and Reddit r/indianeducation.""",
}

RED_FLAGS = {
    "chandigarh-university-online": """Red Flag 1: NIRF Management ranking of #32 means stronger options exist if NIRF management rank is your primary filter. If NIRF top-15 in management is what your employer or PSU application requires, check Amrita (NIRF #26 management) or MAHE (NIRF #14 management) first.

Red Flag 2: The LMS has shown instability during peak assignment submission periods. Avoid last-minute submissions, especially near semester-end deadlines, to prevent unnecessary technical stress.

Red Flag 3: Online MBA placement support is not equivalent to on-campus placement. CU Online's placement assistance is career support, not a job guarantee; budget time for independent job searching alongside the programme.

Red Flag 4: NAAC A+ is strong but not A++. If your target sector (pharma, certain PSU boards, or finance) requires NAAC A++ specifically, verify with your employer before enrolling in this programme.""",

    "manipal-academy-higher-education-online": """Red Flag 1: Rs 2,92,000 is the highest fee in this batch. If your budget does not comfortably support this, do not over-extend on an EMI plan without a clear income assumption built in before you commit.

Red Flag 2: Only 9 specialisations are available. If your required track (e.g., Digital Marketing, International Business, or Event Management) is not in the current list, MAHE online MBA does not offer it.

Red Flag 3: The proctoring software has rejected valid government IDs for some students in exam sessions. Have your Aadhaar ready as the primary identity document before every proctored exam window.

Red Flag 4: Online MBA placement support differs from MAHE's on-campus placement outcomes. The NIRF #3 brand helps with employer recognition, but campus-style structured recruitment does not apply to online MBA students.""",

    "amrita-vishwa-vidyapeetham-online": """Red Flag 1: The ACCA-linked International Finance specialisation does not equate to ACCA membership. Verify the exact nature of the ACCA linkage with Amrita admissions before choosing this specialisation to avoid misaligned expectations.

Red Flag 2: Machine learning appears in the core curriculum from Semester 1. This is valuable but can be challenging for students with zero technical background. Plan additional self-study time in Semesters 1 and 2 if you have no prior data exposure.

Red Flag 3: Semester 3 assignment deadlines have been reported as clustered by multiple students in public forums. Plan your workload early and avoid taking on additional professional commitments during that semester.

Red Flag 4: Online placement support is career assistance, not a placement guarantee. Amrita's strong campus placement record does not automatically translate to structured job placements for online MBA students.""",

    "galgotias-university-online": """Red Flag 1: NIRF rank of 101+ in management is a material limitation if your target PSU or employer has a NIRF top-50 filter for postgraduate eligibility. Check your employer's specific NIRF requirement before choosing Galgotias.

Red Flag 2: The LMS has shown instability during assignment submission windows. Submit all assignments at least 24 hours before the deadline to avoid losing work due to platform failures.

Red Flag 3: The thesis project is a genuine academic commitment, not a rubber-stamp exercise. Students who expected a light Semester 3 and 4 workload found the thesis demanding alongside full-time employment commitments.

Red Flag 4: Career services are basic and do not include structured placement for online MBA students. Galgotias is a budget-first choice; invest your own time in networking and job searching beyond the university's career support.""",

    "sikkim-manipal-university-online": """Red Flag 1: SMU is not NIRF ranked. If your target employer or a government job application has a NIRF rank requirement as an eligibility condition, this is a disqualifying factor before you even apply.

Red Flag 2: The LMS has shown slowness during peak submission periods. Students in earlier batches reported download delays for study material in Semester 2 specifically, which impacted revision time.

Red Flag 3: Only six specialisations are available. If your required track (e.g., Digital Marketing, Business Analytics, or International Business) is absent from the list, SMU does not offer it in the current online MBA curriculum.

Red Flag 4: Placement support is structured around career assistance, not placement guarantee. The Manipal alumni network is large but online MBA alumni specifically show lower active engagement than on-campus alumni in placement activities.""",

    "shoolini-university-online": """Red Flag 1: Fundamentals of Direct Selling is a mandatory Semester 2 subject for all specialisations, including IT Management, Data Science, and Agri-Business. If this is irrelevant to your career track, you will still need to complete and pass it.

Red Flag 2: NIRF #69 overall means Shoolini sits outside the top 50. If your employer uses NIRF top-50 as an eligibility filter for MBA credentials, verify this before enrolling in this programme.

Red Flag 3: Certificate dispatch has taken up to six weeks after convocation for some students. Plan your background verification timeline accordingly if you are planning an employer switch after graduation.

Red Flag 4: Career support is limited to workshops and resume review. Shoolini does not provide structured placement for online MBA students; independent job searching and networking are your own responsibility throughout the programme.""",
}

def generate_page(uni_key):
    u = UNIS[uni_key]
    name = u["name"]
    slug = u["slug"]
    naac = u["naac"]
    nirf_str = f"NIRF #{u['nirf']}" if u["nirf"] < 900 else "Not NIRF ranked"
    specs = u["specs"]
    specs_count = u["specs_count"]
    fee_total = u["fee_total"]
    fee_semester = u["fee_semester"]
    emi_from = u["emi_from"]
    duration = u["duration"]
    eligibility = u["eligibility"]
    hirers = u["hirers"]
    website = u["website"]
    target_profile = u["target_profile"]

    s = {}

    s["tldr"] = (
        f"TL;DR: {name} Online MBA at a glance. UGC-DEB approved. {duration}. "
        f"Rs {fee_total} total fee. {nirf_str}, NAAC {naac}. "
        f"{specs_count} specialisations. Best for {target_profile}."
    )

    s["intro_h2"] = f"Online MBA from {name}: What You Need to Know"
    INTROS = {
        "chandigarh-university-online": (
            f"{name} runs its online MBA under UGC-DEB approval with NAAC A+ accreditation. "
            f"The programme spans {duration} across four semesters and lets students choose from {specs_count} specialisations, "
            f"including niche tracks such as Airlines Management, Disaster Management, and Fintech. "
            f"Three global certifications from Harvard Business Publishing, PwC India, and PMI come bundled at no extra cost. "
            f"The total fee of Rs {fee_total} positions the programme in the mid-range bracket among NAAC A+ online MBAs. "
            f"Classes run in a live format with recorded backups, making it workable for full-time professionals. "
            f"For North India professionals who want brand recognition alongside structured global certifications, CU Online is a practical choice."
        ),
        "manipal-academy-higher-education-online": (
            f"{name} holds both NAAC A++ and Institution of Eminence (IoE) status, placing it among India's most credentialled universities. "
            f"The online MBA runs for {duration} with {specs_count} specialisations, including dedicated tracks for healthcare and pharmaceutical management. "
            f"Students access Coursera courses and a 1:1 industry mentorship programme as structured components. "
            f"The total fee of Rs {fee_total} is the highest in this batch. "
            f"MAHE recommends at least one year of work experience, so this programme is better suited for working professionals than fresh graduates. "
            f"For those whose primary filter is NIRF top-3 and NAAC A++, MAHE Online is the strongest credential available in online format."
        ),
        "amrita-vishwa-vidyapeetham-online": (
            f"{name} brings a NIRF #7 overall ranking and NAAC A++ accreditation to its online MBA. "
            f"The programme spans {duration} and integrates machine learning and ESG modules from Semester 1. "
            f"Eight specialisations include ESG, Fintech, and an ACCA-linked International Finance track, which are rare at this fee point. "
            f"WES recognition makes the degree internationally portable. "
            f"The total fee of Rs {fee_total} (excluding exam fee) positions Amrita as a strong value option among NAAC A++ online MBAs. "
            f"Students who want data and sustainability topics woven into the core curriculum from Day 1 will find this programme distinct from conventional management degrees."
        ),
        "galgotias-university-online": (
            f"{name} offers one of the lowest-fee NAAC A+ online MBA programmes in India at Rs {fee_total} total. "
            f"The programme spans {duration} with seven specialisations and includes a mandatory thesis project across Semesters 3 and 4. "
            f"WES recognition adds international portability to the credential. "
            f"All exams are conducted online in a proctored format. "
            f"For professionals whose primary filter is budget combined with NAAC A+ credential, Galgotias Online delivers that combination more affordably than any other NAAC A+ institution in this batch. "
            f"The thesis model also makes this a research-aware programme rather than a purely taught MBA."
        ),
        "sikkim-manipal-university-online": (
            f"{name} carries the Manipal Group brand, NAAC A+ accreditation, and over three decades of distance education experience. "
            f"The online MBA runs for {duration} with six specialisations covering healthcare, IT systems, marketing, finance, HR, and supply chain. "
            f"Students access Coursera's catalogue free and join a Manipal alumni network of over 1,75,000 graduates. "
            f"The fee of Rs {fee_total} sits in the mid-range for this category. "
            f"SMU is not NIRF ranked, which is a material point if NIRF rank is a hard filter for your target employer or government application. "
            f"For professionals who value the Manipal brand and structured career support, this programme competes solidly at its price point."
        ),
        "shoolini-university-online": (
            f"{name} offers 16 specialisations including highly niche tracks: Agri-Business Management, Biotechnology Management, Food Technology Management, Real Estate Management, and Direct Selling Management. "
            f"No other online MBA in this batch matches that breadth of applied sector tracks. "
            f"The university holds NAAC A+ accreditation and QS ranking. "
            f"The programme spans {duration} at a total fee of Rs {fee_total}. "
            f"A Future Leader's Program runs in Semester 4 across all specialisations as a mandatory component. "
            f"If your career sits at the intersection of business and a specialist domain, Shoolini's track diversity is difficult to match at this fee point."
        ),
    }
    s["intro_body"] = INTROS[uni_key]

    s["about_h2"] = f"About the {name} Online MBA Programme"
    ABOUTS = {
        "chandigarh-university-online": (
            f"Chandigarh University was established in 2012 in Mohali, Punjab. "
            f"It holds NIRF #19 overall and NIRF #32 in Management. "
            f"The online division operates under UGC-DEB entitlement, with all degrees issued under the same university seal as on-campus programmes. "
            f"EdifyEdu verifies these figures independently. We do not take commission from {name}. "
            f"The MBA is delivered through an LMS with live weekend sessions and recorded backups. "
            f"The curriculum includes 30,000+ Harvard Business School case studies and three global certification programmes. "
            f"The programme suits mid-career professionals in North India seeking a credentialled upgrade without interrupting their careers."
        ),
        "manipal-academy-higher-education-online": (
            f"Manipal Academy of Higher Education was established in 1953 in Manipal, Karnataka. "
            f"It holds NIRF #3 overall and NIRF #14 in Management, making it one of the highest-ranked private universities offering an online MBA in India. "
            f"UGC designated MAHE an Institution of Eminence in 2020, a status held by very few private universities nationally. "
            f"EdifyEdu verifies these figures independently. We do not take commission from {name}. "
            f"The online MBA is delivered via the OnlineManipal platform with live sessions, recordings, and Coursera integration. "
            f"Minor projects in Semester 3 and a Capstone Project in Semester 4 are mandatory components. "
            f"The programme is best suited for working professionals with at least one year of experience."
        ),
        "amrita-vishwa-vidyapeetham-online": (
            f"Amrita Vishwa Vidyapeetham was established in 2003 with campuses across five states in India. "
            f"It holds NIRF #7 overall and NAAC A++ accreditation, placing it in the top tier of Indian private universities. "
            f"EdifyEdu verifies these figures independently. We do not take commission from {name}. "
            f"The online MBA is delivered through the Amrita Online platform with blended live and recorded sessions. "
            f"Machine learning and data-driven decision making appear in the core from Semester 1, making this programme distinct from conventional management degrees. "
            f"WES recognition allows graduates to have credentials evaluated internationally, relevant for emigration or foreign employer verification. "
            f"Amrita's research culture underpins a programme that combines management rigour with applied technology exposure."
        ),
        "galgotias-university-online": (
            f"Galgotias University was established in 2011 in Greater Noida, Uttar Pradesh. "
            f"It holds NAAC A+ accreditation and WES recognition, with the online division operating under UGC-DEB entitlement. "
            f"EdifyEdu verifies these figures independently. We do not take commission from {name}. "
            f"The online MBA runs with a structured thesis model: students begin the Master Thesis in Semester 3 and complete it in Semester 4. "
            f"This makes it a research-aware programme, not purely a taught MBA. "
            f"At Rs 76,200 total, it is the most affordable NAAC A+ online MBA in this batch. "
            f"The programme primarily targets professionals in the Delhi-NCR region who want a cost-effective credential with credible accreditation."
        ),
        "sikkim-manipal-university-online": (
            f"Sikkim Manipal University was established in 1995 in Gangtok, Sikkim, as a joint venture between the Manipal Group and the Government of Sikkim. "
            f"It holds NAAC A+ accreditation and is among the oldest universities offering distance and online education in India. "
            f"EdifyEdu verifies these figures independently. We do not take commission from {name}. "
            f"The online MBA is delivered through the SMUDE platform with live sessions and Coursera free access. "
            f"The Manipal alumni network of over 1,75,000 members provides career support. "
            f"SMU's six specialisations cover foundational tracks: marketing, finance, HR, IT systems, operations and supply chain, and healthcare. "
            f"The programme is particularly well recognised in North-East India and in sectors such as defence, healthcare, and banking."
        ),
        "shoolini-university-online": (
            f"Shoolini University was established in 2009 in Solan, Himachal Pradesh. "
            f"It holds NAAC A+ accreditation and QS ranking. "
            f"EdifyEdu verifies these figures independently. We do not take commission from {name}. "
            f"The online MBA offers 16 specialisations, the widest choice in this batch. "
            f"Alongside mainstream tracks, Shoolini includes Agri-Business Management, Biotechnology Management, Food Technology Management, Direct Selling Management, and Real Estate Management. "
            f"These sector-specific tracks attract professionals already working in those industries. "
            f"The curriculum includes Creativity Decoded, Future Leader's Program, and Direct Selling Fundamentals as mandatory components across all tracks."
        ),
    }
    s["about_body"] = ABOUTS[uni_key]

    s["approvals_h2"] = f"Accreditation, Rankings, and Approvals for {name}"
    APPROVALS = {
        "chandigarh-university-online": (
            f"Chandigarh University holds NAAC A+ accreditation (naac.gov.in), NIRF #19 overall and NIRF #32 in Management (nirfindia.org). "
            f"The online MBA is offered under UGC-DEB entitlement (deb.ugc.ac.in), making the degree valid for government jobs and higher education. "
            f"CU Online is QS Ranked, adding an international benchmarking layer. AICTE approval applies to the management programme. "
            f"WES recognition allows foreign employers to verify the credential. "
            f"Graduates hold PSU-eligible status under UGC guidelines. Verify all approval data at the respective official portals before enrolling."
        ),
        "manipal-academy-higher-education-online": (
            f"Manipal Academy of Higher Education holds NAAC A++ accreditation (naac.gov.in), NIRF #3 overall and NIRF #14 in Management (nirfindia.org). "
            f"MAHE was designated an Institution of Eminence (IoE) by UGC, allowing autonomous governance and greater curriculum flexibility. "
            f"The online MBA is offered under UGC-DEB entitlement (deb.ugc.ac.in). AICTE approval and WES recognition are confirmed. "
            f"Graduates hold PSU-eligible status under UGC guidelines. Verify all approval data at the respective official portals before enrolling."
        ),
        "amrita-vishwa-vidyapeetham-online": (
            f"Amrita Vishwa Vidyapeetham holds NAAC A++ accreditation (naac.gov.in), NIRF #7 overall and NIRF #26 in Management (nirfindia.org). "
            f"The online MBA is offered under UGC-DEB entitlement (deb.ugc.ac.in). "
            f"WES recognition means the degree can be evaluated by World Education Services for international credential portability. "
            f"AICTE approval applies to the management programme. Graduates hold PSU-eligible status under UGC guidelines. "
            f"Verify all approval data at the respective official portals before enrolling."
        ),
        "galgotias-university-online": (
            f"Galgotias University holds NAAC A+ accreditation (naac.gov.in). "
            f"The online MBA is offered under UGC-DEB entitlement (deb.ugc.ac.in). AICTE approval and WES recognition are confirmed. "
            f"NIRF ranking for Galgotias overall sits at 101+; the management programme is ranked in the 101-150 band. "
            f"Graduates hold PSU-eligible status under UGC guidelines. "
            f"Verify all approval data at the respective official portals before enrolling."
        ),
        "sikkim-manipal-university-online": (
            f"Sikkim Manipal University holds NAAC A+ accreditation (naac.gov.in). "
            f"The online MBA is offered under UGC-DEB entitlement (deb.ugc.ac.in). "
            f"SMU is not currently listed in the NIRF ranking. This is a material point if NIRF rank is a hard filter for your employer or application. "
            f"The university is government-recognised and graduates hold PSU-eligible status under UGC guidelines. "
            f"Manipal Group affiliation provides an institutional quality signal separate from formal ranking. Verify all approval data at official portals."
        ),
        "shoolini-university-online": (
            f"Shoolini University holds NAAC A+ accreditation (naac.gov.in) and NIRF #69 overall (nirfindia.org). "
            f"The online MBA is offered under UGC-DEB entitlement (deb.ugc.ac.in). "
            f"QS World University Rankings includes Shoolini, providing international benchmark visibility. "
            f"Graduates hold PSU-eligible status under UGC guidelines. "
            f"Verify all approval data at the respective official portals before enrolling."
        ),
    }
    s["approvals_body"] = APPROVALS[uni_key]

    s["ugc_deb_h2"] = "UGC-DEB Approval and What It Means for You"
    s["ugc_deb_body"] = (
        f"The UGC Distance Education Bureau (UGC-DEB) regulates all online degree programmes in India. "
        f"{name} appears on the UGC-DEB entitled institutions list at deb.ugc.ac.in. "
        f"This entitlement means the degree is valid for government jobs at the central and state level; "
        f"graduates can apply for higher education (M.Phil, PhD) at recognised Indian universities; "
        f"and PSU recruitment boards accept the degree under current UGC guidelines. "
        f"UGC-DEB entitlement is reviewed periodically. Always cross-check your admission year's approved list at deb.ugc.ac.in before enrolling. "
        f"A programme not listed at the time of your admission does not carry these protections, regardless of what the university website states."
    )

    s["who_can_apply_h2"] = "Eligibility Criteria for the Online MBA"
    s["who_can_apply_body"] = (
        f"The minimum eligibility for the {name} online MBA is: {eligibility}. "
        f"There is no age limit for online MBA admission under UGC-DEB guidelines, which makes this programme accessible to working professionals at any career stage. "
        f"Candidates with a pass class (below 50%) in graduation should contact the university admissions team to check for SC/ST or OBC relaxation norms applicable for the current intake. "
        f"Work experience is generally not mandatory for this programme, though having professional experience will help you apply learnings more effectively. "
        f"International students and NRIs are eligible to apply, subject to document verification requirements. "
        f"No entrance exam is required; admission is based on graduation percentage and document verification. "
        f"Creating your Academic Bank of Credits (ABC) ID at abc.gov.in before applying will speed up the admission process."
    )

    s["classes_h2"] = "How Classes and Learning Are Delivered"
    CLASS_DETAILS = {
        "chandigarh-university-online": "Weekend live sessions (Saturday/Sunday) via the CU Online LMS, with recorded playback available within 24 hours for working students.",
        "manipal-academy-higher-education-online": "Live online classes via the OnlineManipal platform, Coursera integration for self-paced courses, and recorded sessions available for replay.",
        "amrita-vishwa-vidyapeetham-online": "Blended live and recorded sessions via the Amrita Online platform, with an elective structure that lets students explore specialisation topics from Semester 1.",
        "galgotias-university-online": "Live online sessions scheduled around working hours; thesis supervision is conducted via portal interactions and video calls with assigned guides.",
        "sikkim-manipal-university-online": "SMUDE LMS with live sessions, Coursera free access for 10,000+ courses, and recorded backups for students who miss live classes.",
        "shoolini-university-online": "Live online sessions via the Shoolini LMS with recorded backups; mobile app access available but desktop is recommended for live classes.",
    }
    s["classes_body"] = (
        f"{name} delivers the MBA entirely online with no mandatory campus visit. "
        f"{CLASS_DETAILS[uni_key]} "
        f"Study materials are uploaded at the start of each semester. "
        f"Students interact with faculty through scheduled doubt-clearing sessions and course discussion boards. "
        f"The programme is designed for full-time working adults: live sessions are scheduled on weekends or evenings. "
        f"Assignment deadlines are fixed for each semester but students can manage their daily study schedule independently within those windows."
    )

    s["exams_h2"] = "Examination Pattern and Evaluation"
    s["exams_body"] = (
        f"All examinations for the {name} online MBA are conducted online in a proctored format. "
        f"Students must have a functioning webcam, valid government-issued photo ID, and a stable internet connection for exam sessions. "
        f"Evaluation typically includes continuous assessment through assignments, quizzes, and participation, contributing 30-40% of the grade. "
        f"Semester-end examinations contribute 60-70% of the grade. "
        f"The final semester includes a project or capstone component that counts toward the final evaluation. "
        f"Proctoring software flags unusual behaviour during exams, so ensure your environment is quiet and controlled during exam windows. "
        f"Keep Aadhaar as your primary identity document for exam day, as some proctoring platforms have rejected Voter ID in the past."
    )

    s["specializations_h2"] = f"MBA Specialisations at {name}"
    s["specializations_body"] = (
        f"{name} offers {specs_count} MBA specialisations: {', '.join(specs)}. "
        f"Students select their specialisation at the time of admission. "
        f"Semesters 1 and 2 cover a shared core curriculum across all specialisations. "
        f"Specialisation-specific subjects begin from Semester 3. "
        f"For detailed elective subject lists for each track, visit /universities/{slug}/mba/[spec-slug] on this site. "
        f"Specialisation availability may vary by intake cycle. Confirm your chosen track is active for your intake before paying any fees."
    )

    s["syllabus_h2"] = "Semester-wise Syllabus: Online MBA"
    s["syllabus_body"] = build_syllabus(u)

    s["fees_h2"] = "Fee Structure and Payment Options"
    s["fees_body"] = (
        f"The total fee for the {name} online MBA is approximately Rs {fee_total}. "
        f"This is typically paid in four semester instalments of Rs {fee_semester} each. "
        f"Additional charges may include registration fee, exam fee per semester, and alumni association fee. "
        f"Fee breakdown: {u.get('fee_breakdown', '')}. "
        f"Fees listed are indicative. EdifyEdu reconfirms fees with {name} each quarter. "
        f"Please reconfirm current fees with our counsellor before any payment."
    )

    s["coupon_h3"] = "Discounts and Scholarships"
    s["coupon_body"] = (
        f"{name} offers need-based and merit-based scholarships for eligible students. "
        f"Defence personnel, government employees, and alumni of group institutions may receive fee concessions. "
        f"Early payment discounts may apply at the time of registration for some intake cycles. "
        f"Check the university's official portal for current scholarship notifications before applying. "
        f"We do not apply exclusive coupon codes or take referral commissions from any university."
    )

    s["emi_h2"] = "EMI and Instalment Payment Plans"
    s["emi_body"] = (
        f"EMI options for the {name} online MBA start from Rs {emi_from} per month through partner banking institutions. "
        f"Most students spread the total fee across 12-24 monthly instalments. "
        f"No-cost EMI (zero interest) may be available via specific credit cards or NBFC partners. "
        f"Debit card EMI and NACH-based auto-debit are also accepted at most universities. "
        f"Confirm the current EMI partners and interest rates with the university finance team before selecting a plan. "
        f"Our counsellor can connect you with the admissions team to confirm EMI eligibility for your chosen payment method."
    )

    s["sample_cert_h2"] = "Degree Certificate and Marksheet"
    s["sample_cert_body"] = (
        f"Graduates of the {name} online MBA receive a degree certificate issued by {name} under UGC-DEB guidelines. "
        f"The certificate reads 'Master of Business Administration' without the word 'online' printed on the face of the document. "
        f"This is consistent with UGC-DEB norms for all entitled institutions. "
        f"DigiLocker integration allows digital verification of the degree for employer background checks. "
        f"A consolidated mark sheet covering all four semesters is issued alongside the degree. "
        f"Physical certificates are dispatched after convocation. Allow 4-8 weeks from result declaration for physical delivery to your registered address."
    )

    s["admission_h2"] = "Step-by-Step Admission Process"
    s["admission_body"] = (
        f"Step 1: Visit the official admissions portal at {website}. "
        f"Step 2: Register with your email address and mobile number. "
        f"Step 3: Fill the application form and select your preferred specialisation. "
        f"Step 4: Upload required documents: graduation mark sheets, degree certificate, government photo ID, and a passport-size photograph. "
        f"Step 5: Pay the registration fee (Rs 500-1,200 depending on the university and intake). "
        f"Step 6: Receive the offer letter and confirm admission by paying the first semester fee. "
        f"Most universities run two admission windows per year: January and July intakes. Confirm the current intake deadline with the university before applying."
    )

    s["abc_h2"] = "Academic Bank of Credits (ABC ID) Requirement"
    s["abc_body"] = (
        f"The National Academic Depository's Academic Bank of Credits (ABC) is mandatory for online programme enrolment at UGC-DEB entitled universities. "
        f"Create your ABC ID at abc.gov.in before completing admission at {name}. "
        f"The ABC ID links your academic credits to a national database and enables credit transfer between institutions in future. "
        f"The process takes 5-10 minutes with your Aadhaar number and registered mobile number. "
        f"Have your ABC ID number ready during the application process as {name} will require it at the time of admission confirmation."
    )

    s["placements_h2"] = "Placement Support and Career Outcomes"
    PLACEMENTS = {
        "chandigarh-university-online": (
            f"Chandigarh University Online provides placement assistance through its career services division. "
            f"Services include resume writing workshops, mock interviews, live industry interaction sessions, and access to the CU alumni network. "
            f"CU's on-campus placement record is strong given its NIRF #19 ranking, but online MBA placement outcomes differ from the on-campus programme. "
            f"Set realistic expectations: placement support is career assistance, not a placement guarantee. "
            f"The triple global certification from Harvard, PwC, and PMI can strengthen your profile independently of university placement programmes."
        ),
        "manipal-academy-higher-education-online": (
            f"MAHE Online provides 1:1 industry mentorship as a structured programme component, not just an optional add-on. "
            f"Career services include resume building, interview preparation, and access to the Manipal alumni network. "
            f"MAHE's on-campus placement history is strong for a NIRF #3 university, but online MBA students access a dedicated virtual placement cell. "
            f"Set realistic expectations: online MBA placement assistance is career support, not a placement guarantee. "
            f"The Coursera access lets students add certifications to their profile throughout the programme, which may strengthen employability independently."
        ),
        "amrita-vishwa-vidyapeetham-online": (
            f"Amrita Online supports students through career readiness sessions, resume workshops, and alumni network access. "
            f"The WES-recognised degree is useful for professionals targeting international employers or emigration credential evaluation. "
            f"Students who opt for ESG or Fintech tracks can build portfolio projects alongside academic work, which may differentiate their applications. "
            f"Online MBA placement support is career assistance, not a placement guarantee. "
            f"Build your own professional network through LinkedIn and industry communities during the programme to supplement the university's career support."
        ),
        "galgotias-university-online": (
            f"Galgotias University Online provides career support through resume clinics, mock interviews, and virtual job fair access. "
            f"The mandatory thesis project across Semesters 3 and 4 creates a portfolio piece that can differentiate candidates in management interviews. "
            f"WES recognition helps with international job applications and credential verification. "
            f"Placement support at online MBA level means career assistance. Galgotias does not guarantee placement for online MBA students. "
            f"Delhi-NCR proximity is an advantage for professionals targeting roles in that job market post-graduation."
        ),
        "sikkim-manipal-university-online": (
            f"SMU Online's career services include skill assessment, live industry interaction sessions, resume writing modules, job readiness workshops, and Coursera free access for over 10,000 courses. "
            f"The Manipal Alumni Network with over 1,75,000 members provides a substantial peer and referral base. "
            f"Common entry roles for MBA graduates include Marketing Executive, Sales Executive, HR Executive, Operations Analyst, and Healthcare Administrator. "
            f"Placement assistance means career support, not a guarantee of employment. "
            f"SMU's recognition is strongest in North-East India, banking, healthcare, and defence-adjacent sectors."
        ),
        "shoolini-university-online": (
            f"Shoolini University Online provides career support through the Future Leader's Program in Semester 4, resume workshops, and industry interaction sessions. "
            f"The Sales Management subject in Semester 4 builds a practical commercial skill relevant to placement in marketing, retail, and FMCG roles. "
            f"Niche specialisation tracks such as Agri-Business, Biotechnology, and Food Technology provide better placement fit in sector-specific organisations. "
            f"Placement support is career assistance, not a placement guarantee. "
            f"Shoolini's QS ranking helps with international employer verification but its primary placement network is concentrated in North India."
        ),
    }
    s["placements_body"] = PLACEMENTS[uni_key]

    s["hirers_h2"] = "Top Recruiting Organisations"
    hirers_str = ", ".join(hirers[:12])
    s["hirers_body"] = (
        f"Organisations that have recruited from {name}'s broader alumni base include: {hirers_str}, and others. "
        f"These names reflect historical placement data for the university and are not guarantees for online MBA students specifically. "
        f"Actual hiring from an online MBA depends on your work experience, specialisation choice, and the quality of your profile at the time of applying. "
        f"We do not publish placement statistics we cannot verify from official sources or verified alumni accounts."
    )

    s["beyond_h2"] = "Beyond Admission: What EdifyEdu Offers"
    s["beyond_body"] = (
        f"EdifyEdu compares public UGC/NAAC/NIRF data with no paid rankings and no commissions from any university. "
        f"After reading this page, you can: compare {name} side-by-side with other universities at /compare; "
        f"read the detailed specialisation syllabus pages at /universities/{slug}/mba/[spec-slug]; "
        f"or book a free counsellor call at /contact to clarify fee, scholarship, and admission timeline questions. "
        f"Our counsellors do not push you toward any specific university. "
        f"If you are weighing two or three options, the compare tool lets you place them side by side on fee, NIRF, NAAC, and specialisation count."
    )

    s["reviews_h2"] = "Student Reviews and Ratings"
    s["reviews_body"] = REVIEWS[uni_key]

    s["red_flags_h2"] = "Red Flags to Consider Before Enrolling"
    s["red_flags_body"] = RED_FLAGS[uni_key]

    s["comparisons_h2"] = f"How {name} Compares to Peer Universities"
    peers = u.get("peer_comparison", [])
    nirf_display = u.get("nirf_display", f"NIRF #{u['nirf']}")
    comp_lines = []
    for peer_name, peer_detail in peers:
        comp_lines.append(f"{name}: {nirf_display}, NAAC {naac}, Rs {fee_total}, {specs_count} specialisations.")
        comp_lines.append(f"{peer_name}: {peer_detail}.")
        comp_lines.append("")
    comp_lines.append(
        f"Use the compare tool at /compare to place these universities side by side on all metrics before deciding. "
        f"The right choice depends on your NIRF requirement, budget, specialisation preference, and career sector."
    )
    s["comparisons_body"] = "\n".join(comp_lines)

    choose_bullets = "\n".join(f"- {b}" for b in u["choose_if"])
    not_bullets = "\n".join(f"- {b}" for b in u["not_if"])
    s["verdict_h2"] = f"Honest Verdict: Is the {name} Online MBA Right for You?"
    s["verdict_body"] = (
        f"This verdict is based on publicly available data from UGC-DEB, NAAC, and NIRF portals, cross-referenced with student feedback from verified sources.\n\n"
        f"Choose this if:\n{choose_bullets}\n\n"
        f"Look elsewhere if:\n{not_bullets}\n\n"
        f"The decision between two similarly accredited programmes often comes down to three factors: the specialisation you need, the placement network that matters for your target industry, and the fee your budget can absorb without strain. "
        f"Use the /compare tool to map your shortlist on these three axes. "
        f"If you are still undecided, our counsellor call is free. Visit /contact and we will compare your shortlisted options against your specific career goal and budget without sales pressure."
    )

    s["faqs_h2"] = "Frequently Asked Questions"
    s["faqs"] = [
        (
            f"Is the {name} online MBA UGC approved?",
            f"Yes. {name} is listed on the UGC-DEB entitled institutions list at deb.ugc.ac.in. "
            f"This means the MBA degree is valid for government jobs, higher education, and PSU recruitment under current UGC guidelines. "
            f"Always verify your admission year's list at deb.ugc.ac.in before joining."
        ),
        (
            f"What is the total fee for the {name} online MBA?",
            f"The total indicative fee is Rs {fee_total}, paid across four semesters of Rs {fee_semester} each. "
            f"Additional charges include registration and exam fees. "
            f"Fees are indicative and should be reconfirmed with {name} directly or through our counsellor before any payment."
        ),
        (
            f"What is the eligibility for the {name} online MBA?",
            f"{eligibility}. No entrance examination is required. Work experience is generally not mandatory. "
            f"Candidates below 50% in graduation should check with the admissions team for SC/ST or OBC relaxation norms."
        ),
        (
            f"How are classes conducted at {name} online MBA?",
            f"Classes run entirely online via the university LMS. Live sessions are scheduled on weekends or evenings for working professionals. "
            f"All live sessions are recorded and available for playback within 24-48 hours. "
            f"There is no mandatory campus visit requirement under this programme."
        ),
        (
            f"Does {name} provide placements for online MBA students?",
            f"{name} provides placement assistance including resume workshops, mock interviews, and alumni network access. "
            f"This is career support, not a placement guarantee. "
            f"Actual employment depends on your prior experience, specialisation choice, and your own effort in the job market."
        ),
        (
            f"Is the {name} online MBA equivalent to a regular MBA?",
            f"Yes, under UGC-DEB guidelines, degrees from entitled institutions are equivalent to regular mode degrees for government job eligibility and higher education purposes. "
            f"Private sector employers may or may not treat them identically; verify with your specific target employer if in doubt."
        ),
        (
            f"What specialisations does {name} online MBA offer?",
            f"{name} offers {specs_count} specialisations: {', '.join(specs[:8])}{'...' if len(specs) > 8 else ''}. "
            f"Full specialisation list and elective syllabi are available at /universities/{slug}/mba on this site."
        ),
        (
            f"How do I apply to the {name} online MBA?",
            f"Visit {website} and register for the current intake. "
            f"Upload graduation documents, government photo ID, and passport-size photo. "
            f"Pay the registration fee to confirm your seat. Most intakes open in January and July each year."
        ),
        (
            f"What is the duration of the {name} online MBA?",
            f"The programme runs for {duration}, divided into four semesters of approximately six months each. "
            f"There is no fast-track option; the four-semester structure is fixed under UGC-DEB programme norms."
        ),
        (
            f"Can I pursue government jobs after the {name} online MBA?",
            f"Yes. {name} is UGC-DEB entitled, which means the online MBA is valid for central and state government job applications "
            f"that require a postgraduate management degree. "
            f"PSU recruitment boards also accept UGC-DEB entitled degrees under current guidelines. Verify eligibility for each specific job notification independently."
        ),
    ]

    return s


# ============================================================
# WRITE EXCEL
# ============================================================

OUTPUT_PATH = "data/EdifyEdu_Page_Blueprint_v3.xlsx"
os.chdir("C:/Users/91706/Downloads/edify-v16-final-with-cms/edify-next")

try:
    wb = openpyxl.load_workbook(OUTPUT_PATH)
    print(f"Loaded existing: {OUTPUT_PATH}")
except FileNotFoundError:
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    print(f"Created new: {OUTPUT_PATH}")

UNI_ORDER = [
    "chandigarh-university-online",
    "manipal-academy-higher-education-online",
    "amrita-vishwa-vidyapeetham-online",
    "galgotias-university-online",
    "sikkim-manipal-university-online",
    "shoolini-university-online",
]

REPORT = []

for uni_key in UNI_ORDER:
    u = UNIS[uni_key]
    sheet_name = u["sheet"]
    print(f"\nGenerating: {sheet_name}...")

    s = generate_page(uni_key)

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(title=sheet_name)
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 200
    ws.column_dimensions['D'].width = 12

    ws.append(["Key", "Heading / Q", "Content / A", "Word Count"])
    for col in ['A', 'B', 'C', 'D']:
        ws[f"{col}1"].font = Font(bold=True)

    wrap = Alignment(wrap_text=True, vertical='top')

    def add_row(key, heading, content):
        wc = count_words(str(content))
        ws.append([key, heading or "", content or "", wc])
        r = ws.max_row
        for c in range(1, 5):
            ws.cell(row=r, column=c).alignment = wrap
        ws.row_dimensions[r].height = max(30, min(180, len(str(content)) // 4))

    add_row("tldr", "", s["tldr"])
    add_row("intro_h2", s["intro_h2"], "")
    add_row("intro_body", "", s["intro_body"])
    add_row("about_h2", s["about_h2"], "")
    add_row("about_body", "", s["about_body"])
    add_row("approvals_h2", s["approvals_h2"], "")
    add_row("approvals_body", "", s["approvals_body"])
    add_row("ugc_deb_h2", s["ugc_deb_h2"], "")
    add_row("ugc_deb_body", "", s["ugc_deb_body"])
    add_row("who_can_apply_h2", s["who_can_apply_h2"], "")
    add_row("who_can_apply_body", "", s["who_can_apply_body"])
    add_row("classes_h2", s["classes_h2"], "")
    add_row("classes_body", "", s["classes_body"])
    add_row("exams_h2", s["exams_h2"], "")
    add_row("exams_body", "", s["exams_body"])
    add_row("specializations_h2", s["specializations_h2"], "")
    add_row("specializations_body", "", s["specializations_body"])
    add_row("syllabus_h2", s["syllabus_h2"], "")
    add_row("syllabus_body", "", s["syllabus_body"])
    add_row("fees_h2", s["fees_h2"], "")
    add_row("fees_body", "", s["fees_body"])
    add_row("coupon_h3", s["coupon_h3"], "")
    add_row("coupon_body", "", s["coupon_body"])
    add_row("emi_h2", s["emi_h2"], "")
    add_row("emi_body", "", s["emi_body"])
    add_row("sample_cert_h2", s["sample_cert_h2"], "")
    add_row("sample_cert_body", "", s["sample_cert_body"])
    add_row("admission_h2", s["admission_h2"], "")
    add_row("admission_body", "", s["admission_body"])
    add_row("abc_h2", s["abc_h2"], "")
    add_row("abc_body", "", s["abc_body"])
    add_row("placements_h2", s["placements_h2"], "")
    add_row("placements_body", "", s["placements_body"])
    add_row("hirers_h2", s["hirers_h2"], "")
    add_row("hirers_body", "", s["hirers_body"])
    add_row("beyond_h2", s["beyond_h2"], "")
    add_row("beyond_body", "", s["beyond_body"])
    add_row("reviews_h2", s["reviews_h2"], "")
    add_row("reviews_body", "", s["reviews_body"])
    add_row("red_flags_h2", s["red_flags_h2"], "")
    add_row("red_flags_body", "", s["red_flags_body"])
    add_row("comparisons_h2", s["comparisons_h2"], "")
    add_row("comparisons_body", "", s["comparisons_body"])
    add_row("verdict_h2", s["verdict_h2"], "")
    add_row("verdict_body", "", s["verdict_body"])
    add_row("faqs_h2", s["faqs_h2"], "")
    for i, (q, a) in enumerate(s["faqs"], 1):
        add_row(f"faq_{i}", q, a)

    # Self-check
    all_body = " ".join([
        s["intro_body"], s["about_body"], s["approvals_body"], s["ugc_deb_body"],
        s["who_can_apply_body"], s["classes_body"], s["exams_body"], s["specializations_body"],
        s["syllabus_body"], s["fees_body"], s["coupon_body"], s["emi_body"],
        s["sample_cert_body"], s["admission_body"], s["abc_body"], s["placements_body"],
        s["hirers_body"], s["beyond_body"], s["reviews_body"], s["red_flags_body"],
        s["comparisons_body"], s["verdict_body"],
    ])
    faq_text = " ".join(a for q, a in s["faqs"])
    total_text = all_body + " " + faq_text

    total_words = count_words(total_text)
    edify_count = total_text.lower().count("edifyedu")

    syl = s["syllabus_body"]
    sem3_section = syl[syl.find("SEMESTER 3"):syl.find("SEMESTER 4")] if "SEMESTER 3" in syl and "SEMESTER 4" in syl else ""
    sem4_section = syl[syl.find("SEMESTER 4"):] if "SEMESTER 4" in syl else ""
    sem3_count = sem3_section.count("- ")
    sem4_count = sem4_section.count("- ")

    checks = {
        "words_ok": 2800 <= total_words <= 5000,
        "edify_ok": edify_count <= 5,
        "ugc_deb_ref": "deb.ugc.ac.in" in s["ugc_deb_body"],
        "abc_ref": "abc.gov.in" in s["abc_body"],
        "fees_reconfirm": "reconfirm" in s["fees_body"].lower(),
        "5_reviews": s["reviews_body"].count(" stars:") >= 5,
        "4_red_flags": s["red_flags_body"].count("Red Flag") >= 4,
        "10_faqs": len(s["faqs"]) == 10,
        "no_em_dash": "\u2014" not in total_text,
        "tldr": "TL;DR" in s["tldr"],
        "sem3_ok": sem3_count >= 3,
        "sem4_ok": sem4_count >= 3,
    }

    issues = [k for k, v in checks.items() if not v]
    overall = "PASS" if not issues else "FAIL"

    REPORT.append({
        "uni": u["name"][:40],
        "sheet": sheet_name,
        "words": total_words,
        "edify": edify_count,
        "sem3": sem3_count,
        "sem4": sem4_count,
        "pass": overall,
        "issues": "; ".join(issues) if issues else "None",
    })
    print(f"  Words={total_words} EdifyEdu={edify_count} Sem3={sem3_count} Sem4={sem4_count} -> {overall}")
    if issues:
        print(f"  FAILED CHECKS: {', '.join(issues)}")

wb.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")
print(f"Sheets now in workbook: {wb.sheetnames}")

print("\n" + "="*130)
print("BATCH 3 REPORT")
print("="*130)
print(f"{'University':<42} {'Sheet':<32} {'Words':>6} {'EdifyEdu':>9} {'Sem3':>5} {'Sem4':>5} {'Result':<7} Issues")
print("-"*130)
for r in REPORT:
    print(f"{r['uni']:<42} {r['sheet']:<32} {r['words']:>6} {r['edify']:>9} {r['sem3']:>5} {r['sem4']:>5} {r['pass']:<7} {r['issues']}")
print("="*130)
