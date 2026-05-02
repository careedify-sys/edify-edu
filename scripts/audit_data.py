"""
Edify data audit: cross-check Excel master against Supabase database.
Flags discrepancies in AICTE, NAAC, NIRF, and AACSB records.

Run from edify-next folder:
  python scripts/audit_data.py
"""
import os
import sys
from pathlib import Path
from collections import defaultdict

try:
    from openpyxl import load_workbook
    from dotenv import load_dotenv
    import httpx
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl python-dotenv httpx")
    sys.exit(1)

# Load .env.local from project root
load_dotenv(Path(__file__).parent.parent / ".env.local")

SUPABASE_URL = os.getenv("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    print("ERROR: Missing Supabase credentials in .env.local")
    sys.exit(1)

REST_URL = f"{SUPABASE_URL}/rest/v1"
HEADERS = {
    "apikey": SUPABASE_SERVICE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    "Content-Type": "application/json",
}

# Find Excel file
EXCEL_CANDIDATES = [
    Path(__file__).parent.parent / "UGC_DEB_Online_Programmes_AY2025-26.xlsx",
    Path("UGC_DEB_Online_Programmes_AY2025-26.xlsx"),
    Path.home() / "Downloads" / "UGC_DEB_Online_Programmes_AY2025-26.xlsx",
]
EXCEL_PATH = next((p for p in EXCEL_CANDIDATES if p.exists()), None)

if not EXCEL_PATH:
    print("ERROR: Excel master not found. Tried:")
    for p in EXCEL_CANDIDATES:
        print(f"  - {p}")
    sys.exit(1)

print(f"Excel: {EXCEL_PATH}")
print(f"Supabase: {SUPABASE_URL}")
print()

wb = load_workbook(EXCEL_PATH, data_only=True)

# === Read Excel data ===
print("Reading Excel...")

# Master_Online - programme-level data
ws_m = wb["Master_Online"]
m_headers = [ws_m.cell(1, c).value for c in range(1, ws_m.max_column + 1)]
def col(name):
    return m_headers.index(name) + 1 if name in m_headers else None

uni_col = col("University Name")
aicte_col = col("AICTE Approved")
naac_grade_col = col("NAAC Grade")
naac_cgpa_col = col("NAAC CGPA")
aacsb_status_col = col("AACSB Status")

# Aggregate per university
excel_data = defaultdict(lambda: {
    "aicte_yes_count": 0,
    "aicte_no_count": 0,
    "aicte_na_count": 0,
    "naac_grade": None,
    "naac_cgpa": None,
    "aacsb_status": None,
    "programme_count": 0,
})

for r in range(2, ws_m.max_row + 1):
    uni = ws_m.cell(r, uni_col).value
    if not uni: continue
    uni = uni.strip()
    excel_data[uni]["programme_count"] += 1

    aicte = str(ws_m.cell(r, aicte_col).value or "").strip().lower()
    if aicte == "yes": excel_data[uni]["aicte_yes_count"] += 1
    elif aicte == "no": excel_data[uni]["aicte_no_count"] += 1
    else: excel_data[uni]["aicte_na_count"] += 1

    if not excel_data[uni]["naac_grade"]:
        g = ws_m.cell(r, naac_grade_col).value if naac_grade_col else None
        if g: excel_data[uni]["naac_grade"] = str(g).strip()
    if not excel_data[uni]["naac_cgpa"]:
        c = ws_m.cell(r, naac_cgpa_col).value if naac_cgpa_col else None
        if c: excel_data[uni]["naac_cgpa"] = c
    if not excel_data[uni]["aacsb_status"]:
        a = ws_m.cell(r, aacsb_status_col).value if aacsb_status_col else None
        if a: excel_data[uni]["aacsb_status"] = str(a).strip()

# University_Summary - university-level info
ws_s = wb["University_Summary"]
s_headers = [ws_s.cell(1, c).value for c in range(1, ws_s.max_column + 1)]
def scol(name):
    return s_headers.index(name) + 1 if name in s_headers else None

s_uni = scol("University Name")
s_type = scol("HEI Type")

uni_type_map = {}
for r in range(2, ws_s.max_row + 1):
    uni = ws_s.cell(r, s_uni).value
    typ = ws_s.cell(r, s_type).value
    if uni and typ:
        uni_type_map[uni.strip()] = typ.strip()

print(f"  Found {len(excel_data)} universities in Master_Online")
print(f"  Found {len(uni_type_map)} entries in University_Summary")

# === Read Supabase data (via REST API) ===
print("\nReading Supabase...")
client = httpx.Client(timeout=30)

def sb_select(table, select="*", limit=1000):
    all_data = []
    offset = 0
    while True:
        r = client.get(
            f"{REST_URL}/{table}",
            headers={**HEADERS, "Prefer": "count=exact"},
            params={"select": select, "limit": limit, "offset": offset},
        )
        if r.status_code >= 400:
            print(f"  ERROR reading {table}: {r.status_code} {r.text[:200]}")
            break
        data = r.json()
        all_data.extend(data)
        if len(data) < limit:
            break
        offset += limit
    return all_data

db_unis = sb_select("universities", "id,name,slug,university_type,brand_slug")
db_accreds = sb_select("accreditations", "university_id,body,category,grade,score,rank")

print(f"  {len(db_unis)} universities, {len(db_accreds)} accreditations")

# Build lookup: uni_id -> { 'AICTE': True/False, 'NAAC': {grade, score}, ... }
db_by_uni = defaultdict(lambda: {"AICTE": False, "NAAC": None, "NIRF": [], "AACSB": False, "NBA": False})
for a in db_accreds:
    uid = a["university_id"]
    body = a["body"]
    if body == "AICTE": db_by_uni[uid]["AICTE"] = True
    elif body == "NAAC":
        db_by_uni[uid]["NAAC"] = {"grade": a.get("grade"), "score": a.get("score")}
    elif body == "NIRF":
        db_by_uni[uid]["NIRF"].append({"category": a.get("category"), "rank": a.get("rank")})
    elif body == "AACSB": db_by_uni[uid]["AACSB"] = True
    elif body == "NBA": db_by_uni[uid]["NBA"] = True

# Match by normalized name
def normalize(s):
    return s.lower().strip().replace(" ", "").replace("-", "").replace(",", "") if s else ""

db_uni_by_norm = {}
for u in db_unis:
    db_uni_by_norm[normalize(u["name"])] = u

# === Audit ===
issues = {
    "aicte_missing": [],
    "aicte_unexpected": [],
    "naac_missing_grade": [],
    "naac_missing_in_db": [],
    "no_db_match": [],
}

for uni_name, data in excel_data.items():
    norm = normalize(uni_name)
    db_uni = db_uni_by_norm.get(norm)

    if not db_uni:
        # try fuzzy: match on first 15 normalized chars
        for n, u in db_uni_by_norm.items():
            if n.startswith(norm[:15]) or norm.startswith(n[:15]):
                db_uni = u
                break

    if not db_uni:
        issues["no_db_match"].append(uni_name)
        continue

    db_info = db_by_uni.get(db_uni["id"], {"AICTE": False, "NAAC": None, "NIRF": [], "AACSB": False, "NBA": False})
    uni_type = uni_type_map.get(uni_name, db_uni.get("university_type", ""))
    is_deemed = "deemed" in (uni_type or "").lower()

    # AICTE check
    excel_says_aicte = data["aicte_yes_count"] > 0
    db_has_aicte = db_info["AICTE"]

    if excel_says_aicte and not db_has_aicte:
        issues["aicte_missing"].append({
            "name": uni_name, "type": uni_type,
            "yes_count": data["aicte_yes_count"],
            "total_progs": data["programme_count"],
        })
    elif is_deemed and not db_has_aicte and not excel_says_aicte:
        issues["aicte_missing"].append({
            "name": uni_name, "type": uni_type,
            "yes_count": 0,
            "total_progs": data["programme_count"],
            "regulatory_flag": True,
        })
    elif db_has_aicte and not excel_says_aicte:
        issues["aicte_unexpected"].append({
            "name": uni_name, "type": uni_type,
            "yes_count": data["aicte_yes_count"],
        })

    # NAAC check
    excel_naac_grade = data["naac_grade"]
    db_naac = db_info["NAAC"]

    if excel_naac_grade and not db_naac:
        issues["naac_missing_in_db"].append({
            "name": uni_name, "excel_grade": excel_naac_grade,
            "excel_cgpa": data["naac_cgpa"],
        })
    elif excel_naac_grade and db_naac and not db_naac.get("grade"):
        issues["naac_missing_grade"].append({
            "name": uni_name, "excel_grade": excel_naac_grade,
            "db_score": db_naac.get("score"),
        })

# === Print Report ===
print("\n" + "=" * 70)
print("DATA AUDIT REPORT")
print("=" * 70)

print(f"\n[AICTE] Missing in DB but should have it: {len(issues['aicte_missing'])}")
for issue in issues["aicte_missing"][:30]:
    flag = " [REGULATORY]" if issue.get("regulatory_flag") else ""
    print(f"  - {issue['name']} ({issue['type']}){flag}")
    print(f"    Excel: {issue['yes_count']}/{issue['total_progs']} programmes marked AICTE=Yes")

print(f"\n[AICTE] Unexpected in DB (Excel says all No/N/A): {len(issues['aicte_unexpected'])}")
for issue in issues["aicte_unexpected"][:30]:
    print(f"  - {issue['name']} ({issue['type']})")

print(f"\n[NAAC] Missing in DB entirely: {len(issues['naac_missing_in_db'])}")
for issue in issues["naac_missing_in_db"][:30]:
    print(f"  - {issue['name']}: Excel has {issue['excel_grade']} (CGPA {issue['excel_cgpa']})")

print(f"\n[NAAC] In DB but missing grade field: {len(issues['naac_missing_grade'])}")
for issue in issues["naac_missing_grade"][:30]:
    print(f"  - {issue['name']}: Excel has {issue['excel_grade']}")

print(f"\n[MATCH] Universities in Excel but not found in DB: {len(issues['no_db_match'])}")
for name in issues["no_db_match"][:20]:
    print(f"  - {name}")

print("\n" + "=" * 70)
print("SUMMARY")
print("=" * 70)
total_issues = sum(len(v) for v in issues.values())
print(f"Total issues flagged: {total_issues}")
print(f"  AICTE missing: {len(issues['aicte_missing'])}")
print(f"  AICTE unexpected: {len(issues['aicte_unexpected'])}")
print(f"  NAAC missing entirely: {len(issues['naac_missing_in_db'])}")
print(f"  NAAC missing grade: {len(issues['naac_missing_grade'])}")
print(f"  Name mismatches: {len(issues['no_db_match'])}")

client.close()
