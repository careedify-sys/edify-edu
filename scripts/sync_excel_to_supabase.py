"""
Syncs the master Excel workbook -> Supabase.
Column-name aligned with actual Excel headers.
"""
import os
import re
import sys
from datetime import datetime, date
from pathlib import Path
from typing import Optional

try:
    from dotenv import load_dotenv
    for env_file in ['.env.local', '.env']:
        env_path = Path(env_file)
        if env_path.exists():
            load_dotenv(env_path)
            print(f"Loaded environment from {env_path.absolute()}")
            break
except ImportError:
    print("python-dotenv not installed.")

try:
    from openpyxl import load_workbook
    import httpx
except ImportError:
    print("Missing dependencies. Run: pip install openpyxl httpx python-dotenv")
    sys.exit(1)

SUPABASE_URL = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

if not SUPABASE_URL or not SUPABASE_SERVICE_KEY:
    print("ERROR: Missing Supabase credentials in .env.local")
    sys.exit(1)

REST_URL = f"{SUPABASE_URL}/rest/v1"
HEADERS = {
    "apikey": SUPABASE_SERVICE_KEY,
    "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation",
}

EXCEL_CANDIDATES = [
    os.environ.get("EXCEL_PATH"),
    "./UGC_DEB_Online_Programmes_AY2025-26.xlsx",
    "../UGC_DEB_Online_Programmes_AY2025-26.xlsx",
]
EXCEL_PATH = None
for c in EXCEL_CANDIDATES:
    if c and Path(c).exists():
        EXCEL_PATH = c
        break

if not EXCEL_PATH:
    print("ERROR: Excel file not found.")
    sys.exit(1)

client = httpx.Client(timeout=30)


def sb_upsert(table, rows, on_conflict="slug"):
    all_results = []
    for i in range(0, len(rows), 50):
        chunk = rows[i:i+50]
        r = client.post(
            f"{REST_URL}/{table}",
            headers={**HEADERS, "Prefer": "return=representation,resolution=merge-duplicates"},
            params={"on_conflict": on_conflict},
            json=chunk,
        )
        if r.status_code >= 400:
            print(f"  ERROR upserting {table}: {r.status_code} {r.text[:200]}")
            continue
        all_results.extend(r.json())
    return all_results


def sb_insert(table, rows):
    for i in range(0, len(rows), 100):
        chunk = rows[i:i+100]
        r = client.post(f"{REST_URL}/{table}", headers=HEADERS, json=chunk)
        if r.status_code >= 400:
            print(f"  ERROR inserting {table}: {r.status_code} {r.text[:200]}")


def sb_delete(table, column, value):
    r = client.delete(f"{REST_URL}/{table}", headers=HEADERS, params={column: f"eq.{value}"})
    if r.status_code >= 400:
        print(f"  ERROR deleting {table}: {r.status_code} {r.text[:200]}")


def sb_delete_in(table, column, values):
    """Delete rows where column in (values)."""
    for i in range(0, len(values), 50):
        chunk = values[i:i+50]
        val_str = ",".join(chunk)
        r = client.delete(f"{REST_URL}/{table}", headers=HEADERS, params={column: f"in.({val_str})"})
        if r.status_code >= 400:
            print(f"  ERROR bulk-deleting {table}: {r.status_code} {r.text[:200]}")


def sb_update(table, data, column, value):
    r = client.patch(f"{REST_URL}/{table}", headers=HEADERS, params={column: f"eq.{value}"}, json=data)
    if r.status_code >= 400:
        print(f"  ERROR updating {table}: {r.status_code} {r.text[:200]}")


def slugify(name: str) -> str:
    s = name.lower()
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s).strip("-")
    return f"{s}-online"


def to_iso_date(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, (date, datetime)):
        return value.isoformat()[:10]
    if isinstance(value, str):
        for fmt in ("%d-%b-%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(value, fmt).date().isoformat()
            except ValueError:
                continue
    return None


def safe_float(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def safe_int(v):
    if v is None or v == "":
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def now_iso():
    return datetime.utcnow().isoformat()


def main():
    print(f"\nLoading Excel from {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH, data_only=True)

    sheet_names = wb.sheetnames
    print(f"Sheets: {sheet_names}")

    # ============================================================
    # 1. UNIVERSITIES from University_Summary
    # ============================================================
    ws = wb["University_Summary"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    def col(name):
        try: return headers.index(name) + 1
        except ValueError: return None

    universities_payload = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, col("University Name") or 5).value
        if not name: continue

        state = ws.cell(r, col("State")).value if col("State") else None
        hei_type = ws.cell(r, col("HEI Type")).value if col("HEI Type") else None
        ugc_status = ws.cell(r, col("UGC-DEB Status")).value if col("UGC-DEB Status") else None

        record = {
            "slug": slugify(name),
            "name": name,
            "short_name": None,
            "city": None,
            "state": state,
            "established_year": None,
            "university_type": hei_type,
            "is_ioe": False,
            "ugc_deb_status": "approved" if (str(ugc_status or "").lower() == "approved") else "approved",
            "data_updated_at": now_iso(),
        }
        universities_payload.append(record)

    print(f"\nUpserting {len(universities_payload)} universities...")
    results = sb_upsert("universities", universities_payload, on_conflict="slug")
    name_to_id = {rec["name"]: rec["id"] for rec in results}
    print(f"[OK] {len(name_to_id)} universities synced")

    # ============================================================
    # 2. PROGRAMMES from Master_Online - also collect AICTE per uni
    # ============================================================
    ws_m = wb["Master_Online"]
    m_headers = [ws_m.cell(1, c).value for c in range(1, ws_m.max_column + 1)]

    def mc(name):
        try: return m_headers.index(name) + 1
        except ValueError: return None

    print("\nSyncing programmes...")
    programmes_payload = []
    seen = set()
    aicte_unis = set()
    aacsb_data = {}

    for r in range(2, ws_m.max_row + 1):
        uni_name = ws_m.cell(r, mc("University Name") or 5).value
        prog_name = ws_m.cell(r, mc("Programme Name") or 6).value
        if not uni_name or not prog_name:
            continue
        uni_id = name_to_id.get(uni_name)
        if not uni_id:
            continue

        # AICTE check (programme-level)
        if mc("AICTE Approved"):
            aicte_val = ws_m.cell(r, mc("AICTE Approved")).value
            if aicte_val and str(aicte_val).strip().lower() in ("yes", "approved", "y", "true", "listed"):
                aicte_unis.add(uni_id)

        # AACSB (only set once per uni)
        if mc("AACSB Status") and uni_id not in aacsb_data:
            aacsb_status = ws_m.cell(r, mc("AACSB Status")).value
            if aacsb_status and str(aacsb_status).strip() not in ("", "-", "Not accredited"):
                school = ws_m.cell(r, mc("AACSB Accredited School")).value if mc("AACSB Accredited School") else None
                aacsb_data[uni_id] = (str(aacsb_status).strip(), school)

        # Programme dedupe
        key = (uni_id, prog_name)
        if key in seen:
            continue
        seen.add(key)

        category = ws_m.cell(r, mc("Programme Category")).value if mc("Programme Category") else None
        ugc_end = ws_m.cell(r, mc("Recognition End Date")).value if mc("Recognition End Date") else None
        notice = ws_m.cell(r, mc("UGC Notice")).value if mc("UGC Notice") else None
        aicte_listed_for_prog = False
        if mc("AICTE Approved"):
            v = ws_m.cell(r, mc("AICTE Approved")).value
            aicte_listed_for_prog = bool(v and str(v).strip().lower() in ("yes", "approved", "y", "true", "listed"))

        programmes_payload.append({
            "university_id": uni_id,
            "name": prog_name,
            "short_name": (category or "")[:20],
            "category": category,
            "duration_years": None,
            "ugc_valid_till": to_iso_date(ugc_end),
            "ugc_notice_ref": str(notice) if notice else None,
            "aicte_listed": aicte_listed_for_prog,
            "data_updated_at": now_iso(),
        })

    print(f"  Programmes prepared: {len(programmes_payload)}")
    print(f"  Universities with AICTE: {len(aicte_unis)}")
    print(f"  Universities with AACSB: {len(aacsb_data)}")

    affected_unis = set(p["university_id"] for p in programmes_payload)
    for uni_id in affected_unis:
        sb_delete("programmes", "university_id", uni_id)
    sb_insert("programmes", programmes_payload)

    from collections import Counter
    counts = Counter(p["university_id"] for p in programmes_payload)
    for uni_id, count in counts.items():
        sb_update("universities", {"programme_count": count}, "id", uni_id)

    print(f"[OK] {len(programmes_payload)} programmes synced")

    # ============================================================
    # 3. ACCREDITATIONS - clear all first then rebuild
    # ============================================================
    print("\nSyncing accreditations...")

    # Clear existing for all universities
    all_uni_ids = list(name_to_id.values())
    sb_delete_in("accreditations", "university_id", all_uni_ids)

    accreditations_payload = []
    now = now_iso()

    # NAAC + NBA from University_Summary
    for r in range(2, ws.max_row + 1):
        uni_name = ws.cell(r, col("University Name") or 5).value
        if not uni_name:
            continue
        uni_id = name_to_id.get(uni_name)
        if not uni_id:
            continue

        # NAAC
        if col("NAAC Grade"):
            grade = ws.cell(r, col("NAAC Grade")).value
            if grade and str(grade).strip() not in ("", "-", "Not accredited", "None"):
                accreditations_payload.append({
                    "university_id": uni_id, "body": "NAAC", "category": None,
                    "status": "active", "grade": str(grade).strip(),
                    "score": safe_float(ws.cell(r, col("NAAC CGPA")).value if col("NAAC CGPA") else None),
                    "rank": None, "total_ranked": None, "via_school": None,
                    "cycle": None, "valid_till": None, "notes": None,
                    "data_updated_at": now,
                })

        # NBA
        if col("NBA Courses"):
            nba_count = safe_int(ws.cell(r, col("NBA Courses")).value)
            if nba_count and nba_count > 0:
                nba_status = ws.cell(r, col("NBA Status")).value if col("NBA Status") else None
                accreditations_payload.append({
                    "university_id": uni_id, "body": "NBA", "category": None,
                    "status": "active" if nba_status == "All Valid" else "mixed",
                    "grade": None,
                    "score": float(nba_count),
                    "rank": None, "total_ranked": None, "via_school": None,
                    "cycle": None, "valid_till": None,
                    "notes": f"{nba_count} courses - {nba_status or 'unknown'}",
                    "data_updated_at": now,
                })

    # AICTE - derive from programme-level data
    for uni_id in aicte_unis:
        accreditations_payload.append({
            "university_id": uni_id, "body": "AICTE", "category": None,
            "status": "active",
            "grade": None, "score": None, "rank": None, "total_ranked": None,
            "via_school": None, "cycle": None, "valid_till": None, "notes": None,
            "data_updated_at": now,
        })

    # AACSB
    for uni_id, (status, school) in aacsb_data.items():
        accreditations_payload.append({
            "university_id": uni_id, "body": "AACSB", "category": None,
            "status": "active",
            "grade": None, "score": None, "rank": None, "total_ranked": None,
            "via_school": school, "cycle": None, "valid_till": None, "notes": None,
            "data_updated_at": now,
        })

    # NIRF from Master_Online
    nirf_columns = [
        ("NIRF: Overall", "Overall"), ("NIRF: University", "University"),
        ("NIRF: Research", "Research"), ("NIRF: Engineering", "Engineering"),
        ("NIRF: Management", "Management"), ("NIRF: Pharmacy", "Pharmacy"),
        ("NIRF: Medical", "Medical"), ("NIRF: Dental", "Dental"),
        ("NIRF: Law", "Law"), ("NIRF: Architecture", "Architecture"),
        ("NIRF: Agriculture", "Agriculture"),
    ]

    total_ranked_map = {
        "Overall": 200, "University": 200, "Engineering": 200,
        "Management": 100, "Pharmacy": 100, "Medical": 50,
        "Dental": 40, "Law": 40, "Research": 50,
        "Architecture": 30, "Agriculture": 40,
    }

    nirf_seen = set()
    for r in range(2, ws_m.max_row + 1):
        uni_name = ws_m.cell(r, mc("University Name") or 5).value
        if not uni_name: continue
        uni_id = name_to_id.get(uni_name)
        if not uni_id: continue

        for col_name, cat_label in nirf_columns:
            if mc(col_name):
                rank = safe_int(ws_m.cell(r, mc(col_name)).value)
                if rank:
                    key = (uni_id, cat_label)
                    if key in nirf_seen: continue
                    nirf_seen.add(key)

                    accreditations_payload.append({
                        "university_id": uni_id, "body": "NIRF", "category": cat_label,
                        "status": "active", "rank": rank,
                        "total_ranked": total_ranked_map.get(cat_label),
                        "grade": None, "score": None, "via_school": None,
                        "cycle": None, "valid_till": None, "notes": None,
                        "data_updated_at": now,
                    })

    print(f"  Accreditations prepared: {len(accreditations_payload)}")
    sb_insert("accreditations", accreditations_payload)

    # Print breakdown
    from collections import Counter as Ct
    by_body = Ct(a["body"] for a in accreditations_payload)
    print(f"\nAccreditation breakdown:")
    for body, n in by_body.most_common():
        print(f"  {body}: {n}")

    print(f"\n[OK] Sync complete at {now_iso()}")
    client.close()


if __name__ == "__main__":
    main()
